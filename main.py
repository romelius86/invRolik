import os
import openpyxl
import csv
import requests
import webbrowser # Para abrir WhatsApp y Correo
from datetime import datetime
from fpdf import FPDF

from textual.app import App, ComposeResult
from textual.binding import Binding
from textual.containers import Vertical, Horizontal, VerticalScroll, Grid
from textual.screen import Screen
from textual.widgets import Header, Footer, Static, Button, DataTable, Label, Input, Checkbox, Select

import database

# --- Mapeo de encabezados ---
DB_TO_EXCEL_MAP = {
    'codigo': ['CODIGO', 'CÓDIGO', 'Codigo'],
    'nombre': ['NOMBRE'],
    'fabricante': ['FABRICANTE'],
    'descripcion': ['DESCRIPCIÓN', 'DESCRIPCION'],
    'precio_venta': ['PRECIO VENTA', 'P.Venta articulo', 'P.VENTA'],
    'precio_compra': ['PRECIO COMPRA', 'P.COMPRA'],
    'unidad': ['UNIDAD'],
    'stock': ['STOCK EN ALMACEN', 'STOCK'],
    'fecha_ingreso': ['FECHA INGRESO'],
    'proveedor_nombre': ['PROVEEDOR'],
    'categoria': ['CATEGORIA', 'categoria']
}

# --- Lógica de Importación ---

def import_products_from_excel(filepath):
    if not os.path.exists(filepath):
        return (False, f"Error: No se encontró el archivo:\n{filepath}")

    try:
        workbook = openpyxl.load_workbook(filepath, data_only=True)
        # Intentar con la hoja "productos" o la primera disponible
        sheet_name = "productos" if "productos" in workbook.sheetnames else workbook.sheetnames[0]
        sheet = workbook[sheet_name]
        
        raw_headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        
        column_map = {}
        for db_field, possible_headers in DB_TO_EXCEL_MAP.items():
            for header_text in possible_headers:
                if header_text in raw_headers:
                    column_map[raw_headers.index(header_text)] = db_field
                    break
        
        if 'codigo' not in column_map.values() and 'nombre' not in column_map.values():
            return (False, "Error: No se encontró columna 'CODIGO' o 'NOMBRE' en el Excel.")

        products_to_import = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row): continue
            product_data = {column_map[i]: row[i] for i in column_map if i < len(row)}
            if product_data.get('codigo') or product_data.get('nombre'):
                # Limpia y prepara cada fila antes de guardarla
                prepared = database.add_product_from_flexible_import(product_data)
                products_to_import.append(prepared)
        
        if products_to_import:
            database.bulk_add_products(products_to_import)
            return (True, f"¡Éxito! Se procesaron {len(products_to_import)} productos desde Excel.")
        else:
            return (False, "No se encontraron datos válidos en el archivo Excel.")
    except Exception as e:
        return (False, f"Error Excel: {e}")

def import_products_from_csv(filepath):
    if not os.path.exists(filepath):
        return (False, f"Error: No se encontró el archivo:\n{filepath}")

    detected_encoding = None
    encodings_to_try = ['utf-8-sig', 'latin-1', 'cp1252']
    for encoding in encodings_to_try:
        try:
            with open(filepath, 'r', encoding=encoding) as f: f.read()
            detected_encoding = encoding
            break
        except UnicodeDecodeError: continue

    if not detected_encoding:
        return (False, "Error CSV: No se pudo decodificar el archivo (ni UTF-8 ni ANSI).")

    try:
        with open(filepath, mode='r', encoding=detected_encoding) as f:
            try:
                dialect = csv.Sniffer().sniff(f.read(4096), delimiters=",;")
                f.seek(0)
            except csv.Error:
                dialect = 'excel'
            
            reader = csv.reader(f, dialect)
            
            header_row = None
            data_rows_iterator = None
            for row in reader:
                norm_row = [str(c).strip().upper() for c in row]
                if "CODIGO" in norm_row or "NOMBRE" in norm_row:
                    header_row = row
                    data_rows_iterator = reader # El iterador ahora apunta al resto de las filas
                    break
            
            if not header_row:
                return (False, "Error CSV: No se encontró una fila de encabezado con 'CODIGO' o 'NOMBRE'.")

            raw_headers = header_row
            norm_headers = {h.strip().upper(): h for h in raw_headers}
            column_map = {}
            for db_field, possible_headers in DB_TO_EXCEL_MAP.items():
                for header_text in possible_headers:
                    if header_text.upper() in norm_headers:
                        column_map[norm_headers[header_text.upper()]] = db_field
                        break
            
            products_to_add = []
            for row_list in data_rows_iterator:
                if not any(field for field in row_list): continue
                row_dict = dict(zip(raw_headers, row_list))
                
                product_data_raw = {column_map[h]: row_dict.get(h) for h in column_map if row_dict.get(h) is not None}
                if product_data_raw.get('codigo') or product_data_raw.get('nombre'):
                    # Prepara el diccionario de datos limpios
                    prepared_data = database.add_product_from_flexible_import(product_data_raw)
                    products_to_add.append(prepared_data)
            
            if not products_to_add:
                return (False, "No se encontraron productos válidos para importar en el archivo.")
            
            # Realizar la inserción en bloque
            database.bulk_add_products(products_to_add)
            
            return (True, f"¡Éxito! Se procesaron {len(products_to_add)} productos (codificación: {detected_encoding}).")
    except Exception as e:
        return (False, f"Error al procesar el CSV: {type(e).__name__}: {e}")

def import_products_from_file(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    if ext in ['.xlsx', '.xlsm']:
        return import_products_from_excel(filepath)
    elif ext == '.csv':
        return import_products_from_csv(filepath)
    else:
        return (False, f"Error: Formato '{ext}' no soportado. Usa .xlsx, .xlsm o .csv")

# --- Pantallas de la Aplicación ---

class ImportDialog(Screen):
    def compose(self) -> ComposeResult:
        with Vertical(classes="modal-container"):
            yield Label("IMPORTAR INVENTARIO", id="modal_title")
            yield Label("Pega la ruta del archivo (.csv, .xlsx, .xlsm):")
            yield Input(placeholder="Ej: C:\\documentos\\mi_inventario.csv", id="in_filepath")
            with Horizontal(classes="form-buttons"):
                yield Button("Importar", "success", id="btn_import_start")
                yield Button("Cancelar", "error", id="btn_cancel")

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "btn_import_start":
            filepath = self.query_one("#in_filepath", Input).value.strip('"').strip()
            if not filepath:
                self.app.notify("Por favor, ingresa una ruta.", severity="error")
                return
            
            self.app.notify("Procesando archivo...", severity="information")
            success, message = import_products_from_file(filepath)
            
            if success:
                self.app.notify(message, severity="information", timeout=8)
                self.dismiss(True) # Cierra el diálogo y activa el refresco
            else:
                self.app.notify(message, severity="error", timeout=15)
        
        elif event.button.id == "btn_cancel":
            self.dismiss(False) # Cierra sin refrescar

class LoginScreen(Screen):
    def compose(self) -> ComposeResult:
        with Vertical(classes="modal-container", id="login-container"):
            yield Label("ROLIK - INICIO DE SESIÓN", id="modal_title")
            yield Label("Usuario:")
            yield Input(placeholder="admin / vendedor", id="login_user")
            yield Label("Contraseña:")
            yield Input(placeholder="contraseña", id="login_pass", password=True)
            with Horizontal(classes="form-buttons"):
                yield Button("Ingresar", variant="success", id="btn_login")
                yield Button("Salir", variant="error", id="btn_exit_app")

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "btn_login":
            user_text = self.query_one("#login_user", Input).value
            pw_text = self.query_one("#login_pass", Input).value
            
            user_data = database.authenticate_user(user_text, pw_text) # Devuelve dict o None
            if user_data:
                self.dismiss(user_data) # Pasar el dict completo al callback
            else:
                self.app.notify("Usuario o contraseña incorrectos o inactivos.", severity="error")
        elif event.button.id == "btn_exit_app":
            self.app.exit()

class MainMenu(Screen):
    def compose(self) -> ComposeResult:
        yield Header(show_clock=True)
        user_name = getattr(self.app, "user_name", "Usuario")

        with Vertical(id="menu_container"):
            yield Label(f"SISTEMA ROLIK ERP - {user_name.upper()}", id="title")

            with Grid(id="buttons_grid"):
                if self.app.has_permission('product.view'):
                    yield Button("1. INVENTARIO", id="btn_inventory", variant="primary")

                if self.app.has_permission('purchase_order.view'):
                    yield Button("2. ÓRDENES DE COMPRA", id="btn_purchases", variant="primary")

                # NUEVO BOTÓN DE CLIENTES
                yield Button("3. GESTIÓN DE CLIENTES", id="btn_customers", variant="primary")

                if self.app.has_permission('user.view'):
                    yield Button("USUARIOS", id="btn_users", variant="primary")

                if self.app.has_permission('report.view.sales') or self.app.has_permission('report.view.cash') or self.app.has_permission('report.view.commissions'):
                    yield Button("REPORTES", id="btn_reports", variant="warning")

                if self.app.has_permission('pos.use'):
                    yield Button("4. PUNTO DE VENTA (POS)", id="btn_pos", variant="success")

                if self.app.has_permission('cash.manage'):
                    yield Button("5. GESTIÓN DE CAJA", id="btn_cash", variant="warning")

                yield Button("SALIR", id="btn_exit", variant="error")
        yield Footer()
    def on_button_pressed(self, event: Button.Pressed) -> None:
        screen_map = {
            "btn_inventory": InventoryScreen,
            "btn_purchases": PurchaseOrderListScreen,
            "btn_customers": CustomerManagementScreen, # Nueva pantalla
            "btn_pos": POSScreen,
            "btn_cash": CashScreen,
        }
        if event.button.id in screen_map:
            self.app.push_screen(screen_map[event.button.id]())
        elif event.button.id == "btn_users":
            self.app.push_screen(UserManagementScreen())
        elif event.button.id == "btn_reports":
            self.app.push_screen(ReportsMenuScreen())
        elif event.button.id == "btn_exit":
            self.app.exit()

class CustomerManagementScreen(Screen):
    """Pantalla profesional para el CRUD de clientes."""
    BINDINGS = [
        Binding("escape", "app.pop_screen", "Volver"),
        Binding("n", "add_customer", "Nuevo"),
        Binding("e", "edit_customer", "Editar"),
        Binding("delete", "delete_customer", "Eliminar")
    ]

    def compose(self) -> ComposeResult:
        yield Header()
        yield Label("GESTIÓN INTEGRAL DE CLIENTES", id="section_title")
        with Horizontal(id="customer_search_bar"):
            yield Input(placeholder="Buscar por Nombre o Documento...", id="in_search_cust")
            yield Button("Nuevo (N)", variant="success", id="btn_new_cust")
        
        yield DataTable(id="customers_table")
        
        with Horizontal(id="customer_actions"):
            yield Button("Editar (E)", variant="warning", id="btn_edit_cust")
            yield Button("Eliminar (Supr)", variant="error", id="btn_del_cust")
            yield Button("Volver (ESC)", variant="error", id="btn_back")
        yield Footer()

    def on_mount(self) -> None:
        table = self.query_one(DataTable)
        table.cursor_type = "row"
        table.add_columns("Documento", "Nombre / Razón Social", "Dirección", "Teléfono", "Email")
        self.refresh_customers()

    def on_input_changed(self, event: Input.Changed):
        if event.input.id == "in_search_cust":
            self.refresh_customers(event.value)

    def refresh_customers(self, search=""):
        table = self.query_one(DataTable)
        table.clear()
        for c in database.get_all_customers(search):
            table.add_row(
                str(c['documento']), 
                str(c['nombre']), 
                str(c['direccion'] or '-'), 
                str(c['telefono'] or '-'), 
                str(c['email'] or '-'),
                key=str(c['documento'])
            )

    def action_add_customer(self):
        self.app.push_screen(AddEditCustomerDialog(), lambda success: self.refresh_customers() if success else None)

    def action_edit_customer(self):
        try:
            table = self.query_one(DataTable)
            doc = str(table.get_cell_at((table.cursor_row, 0)))
            client_data = database.buscar_cliente_local(doc)
            if client_data:
                client_data['documento'] = doc
                self.app.push_screen(AddEditCustomerDialog(client_data), lambda success: self.refresh_customers() if success else None)
        except Exception:
            self.app.notify("Seleccione un cliente para editar.", severity="error")

    def action_delete_customer(self):
        try:
            table = self.query_one(DataTable)
            doc = str(table.get_cell_at((table.cursor_row, 0)))
            nombre = str(table.get_cell_at((table.cursor_row, 1)))
            
            # Nota: En una app real pediríamos confirmación. Aquí borramos y notificamos.
            if database.delete_customer(doc):
                self.app.notify(f"Cliente '{nombre}' eliminado.", severity="warning")
                self.refresh_customers()
        except Exception:
            self.app.notify("Seleccione un cliente para eliminar.", severity="error")

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "btn_new_cust": self.action_add_customer()
        elif event.button.id == "btn_edit_cust": self.action_edit_customer()
        elif event.button.id == "btn_del_cust": self.action_delete_customer()
        elif event.button.id == "btn_back": self.app.pop_screen()

class AddEditCustomerDialog(Screen):
    """Diálogo para crear o editar clientes."""
    def __init__(self, client_data=None):
        super().__init__()
        self.client_data = client_data
        self.is_edit = client_data is not None

    def compose(self) -> ComposeResult:
        title = "EDITAR CLIENTE" if self.is_edit else "REGISTRAR NUEVO CLIENTE"
        with Vertical(classes="modal-container", id="cust-form-container"):
            yield Label(title, id="modal_title")
            yield Input(placeholder="DNI o RUC", id="in_doc", disabled=self.is_edit)
            yield Input(placeholder="Nombre o Razón Social", id="in_name")
            yield Input(placeholder="Dirección", id="in_dir")
            yield Input(placeholder="Teléfono", id="in_tel")
            yield Input(placeholder="Correo Electrónico", id="in_email")
            with Horizontal(classes="form-buttons"):
                yield Button("Guardar", variant="success", id="btn_save_cust")
                yield Button("Cancelar", variant="error", id="btn_cancel_cust")

    def on_mount(self):
        if self.is_edit:
            self.query_one("#in_doc").value = self.client_data['documento']
            self.query_one("#in_name").value = self.client_data['nombre']
            self.query_one("#in_dir").value = self.client_data.get('direccion') or ""
            self.query_one("#in_tel").value = self.client_data.get('telefono') or ""
            self.query_one("#in_email").value = self.client_data.get('email') or ""
        self.query_one("#in_name" if self.is_edit else "#in_doc").focus()

    def on_button_pressed(self, event: Button.Pressed):
        if event.button.id == "btn_save_cust":
            doc = self.query_one("#in_doc").value.strip()
            name = self.query_one("#in_name").value.strip()
            if not doc or not name:
                self.app.notify("Documento y Nombre son obligatorios.", severity="error")
                return
            
            data = {
                'documento': doc,
                'nombre': name,
                'direccion': self.query_one("#in_dir").value.strip(),
                'telefono': self.query_one("#in_tel").value.strip(),
                'email': self.query_one("#in_email").value.strip()
            }
            database.add_or_update_customer(data)
            self.dismiss(True)
        else:
            self.dismiss(False)

class UserManagementScreen(Screen):
    BINDINGS = [Binding("escape", "app.pop_screen", "Volver"), ("n", "add_user", "Nuevo"), ("e", "edit_user", "Editar"), ("p", "manage_perms", "Permisos")]

    def compose(self) -> ComposeResult:
        yield Header()
        yield Label("GESTIÓN DE USUARIOS (SOLO ADMINISTRADOR)", id="section_title")
        yield DataTable(id="users_table")
        with Horizontal(id="user_actions"):
            yield Button("Nuevo (N)", "success", id="btn_new_user")
            yield Button("Editar (E)", "primary", id="btn_edit_user")
            yield Button("Permisos (P)", "primary", id="btn_manage_perms")
            yield Button("Activar/Desactivar", "warning", id="btn_toggle_active")
            yield Button("Eliminar", "error", id="btn_delete_user")
            yield Button("Volver (ESC)", "error", id="btn_back")
        yield Footer()

    def on_mount(self) -> None:
        table = self.query_one(DataTable)
        table.cursor_type = "row"
        table.add_columns("ID", "Usuario", "Rol", "Estado", "Desc. Habilitado", "% Desc. Máx.", "Comisión %")
        self.refresh_users()

    def refresh_users(self):
        table = self.query_one(DataTable)
        table.clear()
        for u in database.get_all_users():
            estado = "Activo" if u['is_active'] else "Inactivo"
            desc_habilitado = "Sí" if u['is_discount_enabled'] else "No"
            table.add_row(str(u['id']), u['username'], u['role'], estado, desc_habilitado, f"{u['max_discount_percentage']}%", f"{u['commission_rate']*100:.0f}%")

    def action_add_user(self):
        self.app.push_screen(AddUserDialog(), lambda success: self.refresh_users() if success else None)

    def action_edit_user(self):
        try:
            table = self.query_one(DataTable)
            user_id = int(table.get_cell_at((table.cursor_row, 0)))
            self.app.push_screen(EditUserDialog(user_id=user_id), lambda success: self.refresh_users() if success else None)
        except Exception:
            self.app.notify("Selecciona un usuario para editar.", severity="error")

    def action_manage_perms(self):
        try:
            table = self.query_one(DataTable)
            user_id = int(table.get_cell_at((table.cursor_row, 0)))
            self.app.push_screen(UserPermissionsScreen(user_id=user_id))
        except Exception:
            self.app.notify("Selecciona un usuario para gestionar sus permisos.", severity="error")

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "btn_new_user": self.action_add_user()
        elif event.button.id == "btn_edit_user": self.action_edit_user()
        elif event.button.id == "btn_manage_perms": self.action_manage_perms()
        elif event.button.id == "btn_toggle_active":
            try:
                table = self.query_one(DataTable)
                user_id = int(table.get_cell_at((table.cursor_row, 0)))
                user = database.get_user_by_id(user_id)
                if user['username'] == self.app.user_name:
                    self.app.notify("No puedes desactivarte a ti mismo.", severity="error")
                    return
                
                new_status = not user['is_active']
                database.set_user_active_status(user_id, new_status)
                self.app.notify(f"Usuario '{user['username']}' ahora está {'ACTIVO' if new_status else 'INACTIVO'}.", severity="warning")
                self.refresh_users()
            except Exception:
                self.app.notify("Selecciona un usuario para cambiar su estado.", severity="error")

        elif event.button.id == "btn_delete_user":
            try:
                table = self.query_one(DataTable)
                user_id = int(table.get_cell_at((table.cursor_row, 0)))
                user_name = table.get_cell_at((table.cursor_row, 1))
                if user_name == self.app.user_name:
                    self.app.notify("No puedes eliminarte a ti mismo.", severity="error")
                    return
                database.delete_user(user_id)
                self.app.notify(f"Usuario {user_name} eliminado permanentemente.", severity="warning")
                self.refresh_users()
            except Exception:
                self.app.notify("Selecciona un usuario para eliminar.", severity="error")
        elif event.button.id == "btn_back": self.app.pop_screen()

class UserPermissionsScreen(Screen):
    BINDINGS = [Binding("escape", "app.pop_screen", "Volver")]

    def __init__(self, user_id: int):
        super().__init__()
        self.user_id = user_id
        self.all_permissions = []
        self.user_permissions_ids = set()

    def compose(self) -> ComposeResult:
        yield Header()
        yield Label(f"GESTIONAR PERMISOS", id="section_title")
        with VerticalScroll(id="permissions-container"):
            pass # Se llenará dinámicamente en on_mount
        with Horizontal(classes="form-buttons"):
            yield Button("Guardar Permisos", "success", id="btn_save_permissions")
            yield Button("Cancelar", "error", id="btn_cancel")
        yield Footer()

    def on_mount(self) -> None:
        user = database.get_user_by_id(self.user_id)
        if not user:
            self.app.notify("Usuario no encontrado.", severity="error")
            self.app.pop_screen()
            return
        
        self.query_one("#section_title", Label).update(f"GESTIONAR PERMISOS PARA '{user['username'].upper()}'")
        
        self.all_permissions = database.get_all_permissions()
        user_perms_names = database.get_user_permissions(self.user_id)
        
        container = self.query_one("#permissions-container", VerticalScroll)
        
        for perm in self.all_permissions:
            has_perm = perm['name'] in user_perms_names
            checkbox = Checkbox(f"{perm['name']} - {perm['description']}", value=has_perm, id=f"perm_{perm['id']}")
            container.mount(checkbox)

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "btn_save_permissions":
            permission_ids_to_set = []
            for perm in self.all_permissions:
                checkbox = self.query_one(f"#perm_{perm['id']}", Checkbox)
                if checkbox.value:
                    permission_ids_to_set.append(perm['id'])
            
            try:
                database.update_user_permissions(self.user_id, permission_ids_to_set)
                self.app.notify("Permisos actualizados.", severity="success")
                self.app.pop_screen()
            except Exception as e:
                self.app.notify(f"Error al guardar permisos: {e}", severity="error")
        
        elif event.button.id == "btn_cancel":
            self.app.pop_screen()



class AddUserDialog(Screen):
    def compose(self) -> ComposeResult:
        with Vertical(classes="modal-container"):
            yield Label("CREAR NUEVO USUARIO", id="modal_title")
            yield Label("Nombre:")
            yield Input(placeholder="Usuario", id="new_user_name")
            yield Label("Contraseña:")
            yield Input(placeholder="Password", id="new_user_pass", password=True)
            yield Label("Rol (admin/seller):")
            yield Input(placeholder="admin o seller", id="new_user_role")
            yield Label("Habilitar Descuento (1=Sí, 0=No):")
            yield Input(placeholder="0 o 1", id="new_user_discount_enabled", restrict=r"[01]")
            yield Label("Porcentaje Máximo de Descuento (0-100):")
            yield Input(placeholder="0-100", id="new_user_max_discount", restrict=r"[0-9]*")
            yield Label("Tasa de Comisión (0.00-1.00):")
            yield Input(placeholder="Ej: 0.05 para 5%", id="new_user_commission_rate", restrict=r"[0-9.]*")
            with Horizontal(classes="form-buttons"):
                yield Button("Crear", "success", id="btn_create_user")
                yield Button("Cancelar", "error", id="btn_cancel")

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "btn_create_user":
            name = self.query_one("#new_user_name", Input).value
            pw = self.query_one("#new_user_pass", Input).value
            role = self.query_one("#new_user_role", Input).value.lower().strip()
            discount_enabled_str = self.query_one("#new_user_discount_enabled", Input).value
            max_discount_str = self.query_one("#new_user_max_discount", Input).value
            commission_rate_str = self.query_one("#new_user_commission_rate", Input).value
            
            is_discount_enabled = int(discount_enabled_str) if discount_enabled_str else 0
            max_discount_percentage = int(max_discount_str) if max_discount_str else 0
            commission_rate = float(commission_rate_str) if commission_rate_str else 0.0

            if not (0 <= max_discount_percentage <= 100):
                self.app.notify("El porcentaje de descuento debe ser entre 0 y 100.", severity="error")
                return
            
            if not (0.0 <= commission_rate <= 1.0):
                self.app.notify("La tasa de comisión debe ser entre 0.00 y 1.00.", severity="error")
                return

            if role not in ["admin", "seller"]:
                self.app.notify("Rol inválido. Usa 'admin' o 'seller'.", severity="error")
                return

            if name and pw and role:
                if database.add_user(name, pw, role, is_discount_enabled, max_discount_percentage, commission_rate):
                    self.app.notify("Usuario creado.", severity="success")
                    self.dismiss(True)
                else: self.app.notify("El usuario ya existe.", severity="error")
            else: self.app.notify("Todos los campos son obligatorios.", severity="error")
        elif event.button.id == "btn_cancel": self.dismiss(False)

class EditUserDialog(Screen):
    def __init__(self, user_id: int):
        super().__init__()
        self.user_id = user_id

    def compose(self) -> ComposeResult:
        with Vertical(classes="modal-container"):
            yield Label("EDITAR USUARIO", id="modal_title")
            yield Label("Nombre de Usuario:")
            yield Input(id="edit_user_name")
            yield Label("Rol (admin/seller):")
            yield Input(id="edit_user_role")
            yield Label("Habilitar Descuento (1=Sí, 0=No):")
            yield Input(id="edit_user_discount_enabled", restrict=r"[01]")
            yield Label("Porcentaje Máximo de Descuento (0-100):")
            yield Input(id="edit_user_max_discount", restrict=r"[0-9]*")
            yield Label("Tasa de Comisión (0.00-1.00):")
            yield Input(id="edit_user_commission_rate", restrict=r"[0-9.]*")
            yield Label("Nueva Contraseña (dejar en blanco para no cambiar):")
            yield Input(id="edit_user_pass", password=True, placeholder="Nueva contraseña opcional")
            with Horizontal(classes="form-buttons"):
                yield Button("Guardar Cambios", "success", id="btn_save_user")
                yield Button("Cancelar", "error", id="btn_cancel")

    def on_mount(self) -> None:
        user = database.get_user_by_id(self.user_id)
        if user:
            self.query_one("#edit_user_name", Input).value = user['username']
            self.query_one("#edit_user_role", Input).value = user['role']
            self.query_one("#edit_user_discount_enabled", Input).value = str(user['is_discount_enabled'])
            self.query_one("#edit_user_max_discount", Input).value = str(user['max_discount_percentage'])
            self.query_one("#edit_user_commission_rate", Input).value = str(user['commission_rate'])
        else:
            self.app.notify("Error: No se encontró al usuario.", severity="error")
            self.dismiss(False)

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "btn_save_user":
            name = self.query_one("#edit_user_name", Input).value
            role = self.query_one("#edit_user_role", Input).value.lower().strip()
            is_discount_enabled_str = self.query_one("#edit_user_discount_enabled", Input).value
            max_discount_str = self.query_one("#edit_user_max_discount", Input).value
            commission_rate_str = self.query_one("#edit_user_commission_rate", Input).value
            new_pass = self.query_one("#edit_user_pass", Input).value

            is_discount_enabled = int(is_discount_enabled_str) if is_discount_enabled_str else 0
            max_discount_percentage = int(max_discount_str) if max_discount_str else 0
            commission_rate = float(commission_rate_str) if commission_rate_str else 0.0

            if not (0 <= max_discount_percentage <= 100):
                self.app.notify("El porcentaje de descuento debe ser entre 0 y 100.", severity="error")
                return
            
            if not (0.0 <= commission_rate <= 1.0):
                self.app.notify("La tasa de comisión debe ser entre 0.00 y 1.00.", severity="error")
                return

            if role not in ["admin", "seller"]:
                self.app.notify("Rol inválido. Usa 'admin' o 'seller'.", severity="error")
                return

            if not name or not role:
                self.app.notify("Nombre y Rol son campos obligatorios.", severity="error")
                return

            # Actualizar nombre, rol y configuración de descuento
            if not database.update_user(self.user_id, name, role, is_discount_enabled, max_discount_percentage, commission_rate):
                self.app.notify(f"Error: El nombre de usuario '{name}' ya existe.", severity="error")
                return

            # Actualizar contraseña si se proporcionó una nueva
            if new_pass:
                database.update_user_password(self.user_id, new_pass)
            
            self.app.notify("Usuario actualizado con éxito.", severity="success")
            self.dismiss(True)

        elif event.button.id == "btn_cancel":
            self.dismiss(False)

class InventoryScreen(Screen):
    BINDINGS = [Binding(key, action, desc) for key, action, desc in [
        ("escape", "app.pop_screen", "Volver"), ("n", "add_product", "Nuevo"),
        ("e", "edit_product", "Editar"), ("d", "delete_product", "Eliminar"),
        ("i", "import_excel", "Importar"), ("x", "export_inventory", "Exportar"),
        ("r", "refresh_table", "Refrescar")]]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.search_term = ""
        self.sort_by = "nombre_asc"
    
    def compose(self) -> ComposeResult:
        yield Header()
        yield Label("ROLIK - INVENTARIO DE PRODUCTOS", id="section_title")
        with Horizontal(id="search_sort_bar"):
            yield Input(placeholder="Buscar por código, nombre...", id="search_box")
            yield Button("Código ↓", id="sort_code_desc", variant="primary")
            yield Button("Código ↑", id="sort_code_asc", variant="primary")
            yield Button("Nombre ↓", id="sort_name_desc", variant="primary")
            yield Button("Nombre ↑", id="sort_name_asc", variant="primary")
            yield Button("Stock ↓", id="sort_stock_desc", variant="primary")
            yield Button("Stock ↑", id="sort_stock_asc", variant="primary")

        yield DataTable(id="inventory_table", zebra_stripes=True)
        
        with Horizontal(id="inventory_actions"):
            if self.app.has_permission('product.create'):
                yield Button("Nuevo (N)", "success", id="btn_new")
            if self.app.has_permission('product.edit'):
                yield Button("Editar (E)", "warning", id="btn_edit")
            if self.app.has_permission('product.delete'):
                yield Button("Eliminar (D)", "error", id="btn_delete")
            if self.app.has_permission('product.import'):
                yield Button("Importar (I)", "primary", id="btn_import") 
            yield Button("Exportar (X)", "success", id="btn_export")
            yield Button("Refrescar (R)", "primary", id="btn_refresh")
            yield Button("Volver (ESC)", "error", id="btn_back")
        
        yield Footer()
    
    def on_mount(self) -> None:
        table = self.query_one(DataTable); table.cursor_type = "row"
        table.add_columns(
            "ID", "SKU", "Producto", "Categoría", "Proveedor", "Descripción",
            "Stock Act.", "Stock Mín.", "Stock Máx.", 
            "Costo Prom.", "Precio Venta", "Valor Inv."
        )
        self.refresh_inventory()

    def on_input_changed(self, event: Input.Changed) -> None:
        if event.input.id == "search_box":
            self.search_term = event.value
            self.refresh_inventory()

    def on_product_saved(self, success: bool):
        if success: self.refresh_inventory(); self.app.notify("Producto guardado.", severity="success")
    
    def action_add_product(self): self.app.push_screen(AddEditProductScreen(), self.on_product_saved)
    
    def action_edit_product(self):
        try:
            table = self.query_one(DataTable)
            if table.row_count == 0: raise IndexError()
            codigo = table.get_cell_at((table.cursor_row, 0))
            self.app.push_screen(AddEditProductScreen(product_codigo=codigo), self.on_product_saved)
        except Exception: self.app.notify("Selecciona un producto para editar.", severity="error")

    def action_delete_product(self):
        try:
            table = self.query_one(DataTable)
            if table.row_count == 0: return
            codigo = table.get_cell_at((table.cursor_row, 0))
            database.delete_product(codigo)
            self.refresh_inventory()
            self.app.notify(f"Producto {codigo} eliminado.", severity="warning")
        except Exception: self.app.notify("Error al eliminar el producto.", severity="error")
    
    def refresh_inventory(self):
        try:
            table = self.query_one(DataTable)
            table.clear()
            
            # Pasa los parámetros de búsqueda y orden a la función de la base de datos
            products = database.get_all_products_for_display(
                search_term=self.search_term, 
                sort_by=self.sort_by
            )
            
            for p in products:
                # Formateamos los valores numéricos para que se vean profesionales
                costo_prom = f"S/ {p['costo_promedio']:,.2f}"
                precio_venta = f"S/ {p['precio_venta']:,.2f}"
                valor_inv = f"S/ {p['valor_inventario']:,.2f}"
                
                table.add_row(
                    str(p['id']),
                    str(p['sku']),
                    str(p['nombre']),
                    str(p['categoria'] or "N/A"),
                    str(p['proveedor_nombre'] or "-"),
                    str(p['descripcion'] or "-"),
                    str(p['stock_actual']),
                    str(p['stock_minimo']),
                    str(p['stock_maximo']),
                    costo_prom,
                    precio_venta,
                    valor_inv,
                    key=str(p['sku'])
                )
            
            table.refresh()
                
        except Exception as e:
            self.app.notify(f"Error visual: {e}", severity="error")
    
    def action_refresh_table(self): 
        self.search_term = ""
        self.query_one("#search_box", Input).value = ""
        self.sort_by = "nombre_asc"
        self.refresh_inventory()
    
    def action_import_excel(self):
        self.app.push_screen(ImportDialog(), lambda success: self.refresh_inventory() if success else None)
    
    def action_export_inventory(self):
        """Exporta todo el inventario a un archivo Excel."""
        try:
            products = database.get_all_products_for_display()
            if not products:
                self.app.notify("No hay productos para exportar.", severity="warning")
                return

            os.makedirs("reportes", exist_ok=True)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Inventario ROLIK"

            # Encabezados
            headers = ["CÓDIGO", "NOMBRE", "CATEGORÍA", "FABRICANTE", "DESCRIPCIÓN", "P. VENTA", "P. COMPRA", "UNIDAD", "STOCK", "FECHA INGRESO", "PROVEEDOR"]
            ws.append(headers)

            for p in products:
                ws.append([
                    p['codigo'], p['nombre'], p['categoria'], p['fabricante'], p['descripcion'],
                    p['precio_venta'], p['precio_compra'], p['unidad'], p['stock'], 
                    p['fecha_ingreso'], p['proveedor_nombre']
                ])

            filename = os.path.join("reportes", f"Inventario_ROLIK_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
            wb.save(filename)
            self.app.notify(f"Inventario exportado: {filename}", severity="information")
            if os.name == 'nt': os.startfile(filename)
            
        except Exception as e:
            self.app.notify(f"Error al exportar: {e}", severity="error")

    def on_button_pressed(self, event: Button.Pressed):
        action_map = {
            "btn_new": self.action_add_product, 
            "btn_edit": self.action_edit_product,
            "btn_delete": self.action_delete_product,
            "btn_import": self.action_import_excel, 
            "btn_export": self.action_export_inventory,
            "btn_refresh": self.action_refresh_table,
            "btn_back": self.app.pop_screen
        }
        
        sort_map = {
            "sort_code_asc": "codigo_asc",
            "sort_code_desc": "codigo_desc",
            "sort_name_asc": "nombre_asc",
            "sort_name_desc": "nombre_desc",
            "sort_stock_asc": "stock_asc",
            "sort_stock_desc": "stock_desc",
        }

        if event.button.id in action_map:
            action_map[event.button.id]()
        elif event.button.id in sort_map:
            self.sort_by = sort_map[event.button.id]
            self.refresh_inventory()

class AddEditProductScreen(Screen):
    BINDINGS = [Binding("escape", "app.pop_screen", "Salir")]

    def __init__(self, product_codigo: str | None = None):
        super().__init__(); self.product_codigo, self.is_edit_mode = product_codigo, product_codigo is not None
    
    def compose(self) -> ComposeResult:
        yield Header()
        yield Label("EDITAR PRODUCTO" if self.is_edit_mode else "AÑADIR NUEVO PRODUCTO", id="modal_title")
        
        with VerticalScroll(id="form-body"):
            yield Label("Código:")
            yield Input(id="in_codigo", disabled=self.is_edit_mode)
            yield Label("Nombre:")
            yield Input(id="in_nombre")
            yield Label("Categoría:")
            yield Input(id="in_categoria")
            yield Label("Fabricante:")
            yield Input(id="in_fabricante")
            yield Label("Proveedor:")
            yield Input(id="in_proveedor")
            yield Label("Descripción:")
            yield Input(id="in_descripcion")
            yield Label("P.Venta:")
            yield Input(id="in_precio_venta", restrict=r"[0-9.]*")
            yield Label("P.Compra:")
            yield Input(id="in_precio_compra", restrict=r"[0-9.]*")
            yield Label("Unidad:")
            yield Input(id="in_unidad")
            yield Label("Stock:")
            yield Input(id="in_stock", restrict=r"[0-9]*")
            yield Label("Stock Mínimo:")
            yield Input(id="in_stock_minimo", restrict=r"[0-9]*")

        with Horizontal(classes="form-buttons"):
            yield Button("Guardar cambios", variant="success", id="btn_save")
            yield Button("Salir (sin guardar)", variant="error", id="btn_cancel")
        
        yield Footer()
    
    def on_mount(self):
        if self.is_edit_mode:
            product = database.get_product(self.product_codigo)
            if product:
                p = dict(product)
                for key in ["codigo", "nombre", "categoria", "fabricante", "descripcion", "precio_venta", "precio_compra", "unidad", "stock", "stock_minimo"]: self.query_one(f"#in_{key}", Input).value = str(p.get(key) or '')
                self.query_one("#in_proveedor", Input).value = str(p.get('proveedor_nombre', ''))
    
    def on_button_pressed(self, event: Button.Pressed):
        if event.button.id == "btn_save":
            p_data = {k: self.query_one(f"#in_{k}", Input).value for k in ["codigo", "nombre", "categoria", "fabricante", "descripcion", "precio_venta", "precio_compra", "unidad", "stock", "stock_minimo"]}
            p_data["proveedor_nombre"] = self.query_one("#in_proveedor", Input).value
            if self.is_edit_mode:
                product = database.get_product(self.product_codigo)
                p_data['fecha_ingreso'] = product['fecha_ingreso'] if product else None
            
            if not p_data["codigo"] and not self.is_edit_mode:
                 p_data["codigo"] = f"P-{int(datetime.now().timestamp())}"
            
            if not p_data["nombre"]:
                self.app.notify("El campo 'Nombre' es obligatorio.", severity="error")
                return

            database.add_or_update_product(p_data)
            self.dismiss(True)
        elif event.button.id == "btn_cancel":
            self.dismiss(False)

class POSScreen(Screen):
    BINDINGS = [
        Binding("escape", "app.pop_screen", "Volver"),
        Binding("f2", "finish_sale", "Finalizar"),
        Binding("f4", "clear_sale", "Limpiar")
    ]
    
    def __init__(self): 
        super().__init__()
        self.cart = {} # {codigo: [nombre, cantidad, precio]}
        self.total = 0.0
        self.discount_percentage = 0
    
    def compose(self) -> ComposeResult:
        yield Header()
        with Horizontal():
            with Vertical(id="pos_left"):
                yield Label("ROLIK - PUNTO DE VENTA (POS)", classes="section_title")
                yield Label("Buscar por Código/Nombre:")
                yield Input(id="pos_in_search", placeholder="Escriba código o nombre...")
                yield DataTable(id="pos_table")
            with Vertical(id="pos_right"):
                yield Label("RESUMEN", classes="section_title")
                yield Static("SUBTOTAL: $0.00", id="pos_subtotal_label")
                yield Label("Descuento (%):")
                yield Input("0", id="pos_in_discount", restrict=r"[0-9]*")
                yield Static("TOTAL: $0.00", id="pos_total_label")
                yield Button("Finalizar Pago (F2)", variant="success", id="btn_finish")
                yield Button("Eliminar Item Seleccionado", variant="error", id="btn_delete_item")
                yield Button("Limpiar Todo (F4)", variant="warning", id="btn_clear")
                yield Button("Volver (ESC)", variant="error", id="btn_back")
        yield Footer()
    
    def on_mount(self):
        table = self.query_one("#pos_table")
        table.add_columns("Código", "Producto", "Marca", "Unidad", "Cant.", "Stock", "P.Unit", "Subtotal")
        table.cursor_type = "row"
        self.query_one("#pos_in_search").focus()
        if not database.get_active_session():
            self.app.notify("¡Caja cerrada! Abra la caja para vender.", severity="error")
            self.query_one("#pos_in_search").disabled = True

    def on_input_submitted(self, event: Input.Submitted) -> None:
        if event.input.id == "pos_in_search":
            search_term = event.value.strip()
            if search_term:
                self.handle_search(search_term)
                event.input.value = ""

    def handle_search(self, term):
        # 1. Buscar por código exacto primero (para escáner de barras)
        product = database.get_product(term)
        if product:
            self.prompt_quantity(product)
            return

        # 2. Si no es código exacto, buscar coincidencias por nombre o parte del código
        results = database.get_all_products_for_display(search_term=term)
        
        if len(results) == 0:
            self.app.notify(f"No se encontró nada con '{term}'.", severity="error")
        elif len(results) == 1:
            # Si solo hay uno, lo agregamos directamente
            full_prod = database.get_product(results[0]['codigo'])
            self.prompt_quantity(full_prod)
        else:
            # Si hay varios, mostramos ventana de selección
            self.app.push_screen(ProductSearchDialog(term, results), self.on_product_selected)

    def on_product_selected(self, product_code):
        if product_code:
            product = database.get_product(product_code)
            if product:
                self.prompt_quantity(product)

    def prompt_quantity(self, product):
        """Abre un diálogo para pedir la cantidad antes de añadir al carrito."""
        def check_qty(qty):
            if qty: self.add_to_cart(product, qty)
        self.app.push_screen(QuantityDialog(product), check_qty)

    def add_to_cart(self, product, qty):
        codigo = product['codigo']
        nombre = product['nombre']
        marca = str(product['fabricante'] or "N/A")
        unidad = str(product['unidad'] or "Und")
        stock = product['stock'] or 0
        precio = product['precio_venta'] or 0.0
        
        # Estructura del carrito: [nombre, cantidad, precio, stock, marca, unidad]
        in_cart = self.cart.get(codigo, [None, 0, None, 0, None, None])[1]
        
        if stock < (in_cart + qty):
            self.app.notify(f"Stock insuficiente para '{nombre}'. Disponible: {stock}", severity="error")
            return
        
        if codigo in self.cart:
            self.cart[codigo][1] += qty
        else:
            self.cart[codigo] = [nombre, qty, precio, stock, marca, unidad]
        
        self.refresh_pos_table()

    def refresh_pos_table(self):
        """Actualiza la tabla del carrito y los totales en pantalla."""
        table = self.query_one("#pos_table")
        table.clear()
        subtotal_calc = 0.0
        for codigo, (nombre, qty, price, stock, marca, unidad) in self.cart.items(): 
            item_subtotal = qty * (price or 0)
            subtotal_calc += item_subtotal
            table.add_row(codigo, nombre, marca, unidad, str(qty), str(stock), f"${price or 0:.2f}", f"${item_subtotal:.2f}")
        
        self.query_one("#pos_subtotal_label").update(f"SUBTOTAL: ${subtotal_calc:.2f}")
        total_with_disc = subtotal_calc * (1 - self.discount_percentage / 100)
        self.total = total_with_disc
        self.query_one("#pos_total_label").update(f"TOTAL: ${self.total:.2f}")

    def on_button_pressed(self, event: Button.Pressed):
        if event.button.id == "btn_finish":
            self.action_finish_sale()
        elif event.button.id == "btn_delete_item":
            self.delete_selected_item()
        elif event.button.id == "btn_clear":
            self.action_clear_sale()
        elif event.button.id == "btn_back":
            self.app.pop_screen()

    def delete_selected_item(self):
        table = self.query_one("#pos_table")
        try:
            codigo = table.get_cell_at((table.cursor_row, 0))
            if codigo in self.cart:
                del self.cart[codigo]
                self.refresh_pos_table()
        except Exception:
            self.app.notify("Selecciona un item para eliminar.", severity="error")

    def action_finish_sale(self):
        if not self.cart:
            self.app.notify("Carrito vacío.", severity="warning")
            return
        # AHORA PASAMOS EL CARRITO Y EL TOTAL AL DIÁLOGO
        self.app.push_screen(PaymentDialog(self.cart, self.total), self.on_payment_finished)

    def on_payment_finished(self, payment_result):
        if payment_result:
            session = database.get_active_session()
            items = [(code, data[1], data[2]) for code, data in self.cart.items()]
            try:
                trans_id, correlativo = database.record_sale(session['id'], self.total, items, self.app.user_id, payment_result)
                self.app.notify(f"Venta registrada: {correlativo}", severity="success")
                # Mostrar ticket final con opción de impresión
                self.app.push_screen(FinalReceiptScreen(self.cart.copy(), self.total, payment_result, correlativo))
                self.action_clear_sale()
            except Exception as e:
                self.app.notify(f"Error al grabar: {e}", severity="error")

    def action_clear_sale(self):
        self.cart.clear()
        self.discount_percentage = 0
        self.refresh_pos_table()

class ProductSearchDialog(Screen):
    """Ventana emergente para seleccionar productos cuando la búsqueda devuelve varios resultados."""
    BINDINGS = [Binding("escape", "cancel_search", "Cerrar")]

    def __init__(self, term, results):
        super().__init__()
        self.term = term
        self.results = results

    def compose(self) -> ComposeResult:
        with Vertical(classes="modal-container", id="search-dialog"):
            yield Label(f"RESULTADOS PARA: '{self.term.upper()}'", id="modal_title")
            yield DataTable(id="search_results_table")
            with Horizontal(classes="form-buttons"):
                yield Button("Seleccionar (Enter)", variant="success", id="btn_select_prod")
                yield Button("Cancelar (ESC)", variant="error", id="btn_cancel_search")

    def on_mount(self):
        table = self.query_one("#search_results_table")
        table.add_columns("Código", "Producto", "Marca", "Unidad", "Stock", "Precio")
        table.cursor_type = "row"
        for p in self.results:
            fab = str(p['fabricante'] or "N/A")
            unidad = str(p['unidad'] or "Und")
            table.add_row(
                p['codigo'], p['nombre'], fab, unidad, str(p['stock']), 
                f"${p['precio_venta']:.2f}", 
                key=p['codigo']
            )
        table.focus()

    def on_data_table_row_selected(self, event: DataTable.RowSelected):
        if event.row_key:
            self.dismiss(str(event.row_key.value))

    def action_cancel_search(self):
        self.dismiss(None)

    def on_button_pressed(self, event: Button.Pressed):
        if event.button.id == "btn_select_prod":
            table = self.query_one("#search_results_table", DataTable)
            try:
                if table.cursor_row is not None:
                    # Obtenemos el valor de la celda en la columna 0 (Código)
                    codigo = table.get_cell_at((table.cursor_row, 0))
                    if codigo:
                        self.dismiss(str(codigo))
                else:
                    self.app.notify("Selecciona un producto de la lista.", severity="warning")
            except Exception as e:
                self.app.notify("Error: Seleccione una fila válida primero.", severity="error")
        elif event.button.id == "btn_cancel_search":
            self.dismiss(None)

class QuantityDialog(Screen):
    """Diálogo pequeño para ingresar la cantidad de un producto."""
    def __init__(self, product):
        super().__init__()
        self.product = product

    def compose(self) -> ComposeResult:
        with Vertical(classes="modal-container", id="qty-dialog"):
            yield Label(f"CANTIDAD PARA: {self.product['nombre']}")
            yield Label(f"Stock Disponible: {self.product['stock']}")
            yield Input("1", id="in_qty", restrict=r"[0-9]*")
            with Horizontal(classes="form-buttons"):
                yield Button("Agregar", variant="success", id="btn_add")
                yield Button("Cancelar", variant="error", id="btn_cancel")

    def on_mount(self):
        self.query_one("#in_qty").focus()

    def on_button_pressed(self, event: Button.Pressed):
        if event.button.id == "btn_add":
            try:
                qty = int(self.query_one("#in_qty").value)
                if qty <= 0:
                    self.app.notify("La cantidad debe ser mayor a cero.")
                    return
                self.dismiss(qty)
            except ValueError:
                self.app.notify("Ingrese un número válido.")
        else:
            self.dismiss(None)

class PaymentDialog(Screen):
    def __init__(self, cart, total):
        super().__init__()
        self.cart = cart
        self.total = total
        self.payment_method = "EFECTIVO"
        self.comp_type = "BOLETA"

    def compose(self) -> ComposeResult:
        with Vertical(classes="modal-container", id="pay-dialog-full"):
            yield Label("CONFIRMAR VENTA Y PAGO", id="modal_title")
            
            # PARTE SUPERIOR: CLIENTE Y PAGO (2 Columnas)
            with Grid(id="pay_top_grid"):
                # Columna Cliente
                with Vertical(classes="pay_box"):
                    yield Label("1. DATOS DEL CLIENTE:")
                    with Horizontal(classes="options-row"):
                        yield Button("BOLETA", id="btn_boleta", variant="primary")
                        yield Button("FACTURA", id="btn_factura")
                        yield Button("TICKET", id="btn_ticket")
                    with Horizontal(id="client-search-row"):
                        yield Input(placeholder="DNI/RUC", id="in_client_doc", restrict=r"[0-9]*")
                        yield Button("🔍", id="btn_search_client", variant="primary")
                    yield Input(placeholder="Nombre / Razón Social", id="in_client_name")
                
                # Columna Pago
                with Vertical(classes="pay_box"):
                    yield Label("2. MÉTODO DE PAGO Y COBRO:")
                    with Horizontal(classes="options-row"):
                        yield Button("EFECTIVO", id="btn_efectivo", variant="primary")
                        yield Button("TARJETA", id="btn_tarjeta")
                        yield Button("YAPE/PLIN", id="btn_yape")
                    yield Label("MONTO RECIBIDO:")
                    yield Input(str(self.total), id="in_paid_amount", restrict=r"[0-9.]*")
                    yield Static("VUELTO: S/ 0.00", id="lbl_change")

            # PARTE INFERIOR: REVISIÓN Y EDICIÓN DE PRODUCTOS
            with Vertical(id="pay_bottom_container"):
                yield Label("3. REVISAR Y EDITAR PRODUCTOS:")
                # Buscador extra por si falta algo
                with Horizontal(id="pay_extra_search_row"):
                    yield Input(placeholder="¿Falta algo? Busca por Código o Nombre aquí...", id="in_pay_extra_search")
                
                yield DataTable(id="pay_confirm_table")
                
                with Horizontal(id="pay_edit_controls"):
                    yield Button("➕ Aumentar", id="btn_add_qty", variant="success")
                    yield Button("➖ Disminuir", id="btn_sub_qty", variant="warning")
                    yield Button("❌ Eliminar Item", id="btn_remove_item", variant="error")
                
                yield Static(f"TOTAL A PAGAR: S/ {self.total:.2f}", id="pay_confirm_subtotal")
            
            with Horizontal(classes="form-buttons"):
                yield Button("¡CONFIRMAR VENTA FINAL!", variant="success", id="btn_confirm_pay")
                yield Button("CANCELAR / VOLVER", variant="error", id="btn_cancel_pay")

    def on_input_submitted(self, event: Input.Submitted):
        if event.input.id == "in_pay_extra_search":
            term = event.value.strip()
            if term:
                self.buscar_y_añadir_extra(term)
                event.input.value = ""

    def buscar_y_añadir_extra(self, term):
        """Busca un producto y lo añade al carrito desde la ventana de pago."""
        # 1. Intentar código exacto primero
        product = database.get_product(term)
        if product:
            self.finalizar_añadido_extra(product)
            return

        # 2. Si no, buscar coincidencias similares
        results = database.get_all_products_for_display(search_term=term)
        if len(results) == 0:
            self.app.notify(f"No se encontró nada con '{term}'.", severity="error")
        elif len(results) == 1:
            # Si solo hay uno, lo agregamos directo
            full_prod = database.get_product(results[0]['codigo'])
            self.finalizar_añadido_extra(full_prod)
        else:
            # SI HAY VARIOS, MOSTRAR DIÁLOGO DE SELECCIÓN
            self.app.push_screen(ProductSearchDialog(term, results), self.on_extra_product_selected)

    def on_extra_product_selected(self, product_code):
        """Callback cuando se selecciona un producto del diálogo de búsqueda extra."""
        if product_code:
            product = database.get_product(product_code)
            if product:
                self.finalizar_añadido_extra(product)

    def finalizar_añadido_extra(self, product):
        """Añade físicamente el producto al carrito del diálogo."""
        code = product['codigo']
        if code in self.cart:
            if self.cart[code][1] < product['stock']:
                self.cart[code][1] += 1
            else:
                self.app.notify("Stock insuficiente.", severity="error")
        else:
            # [nombre, qty, price, stock, marca, unidad]
            self.cart[code] = [product['nombre'], 1, product['precio_venta'], product['stock'], product['fabricante'], product['unidad']]
        
        self.refresh_confirm_table()
        self.app.notify(f"Añadido: {product['nombre']}")

    def on_mount(self):
        table = self.query_one("#pay_confirm_table")
        table.add_columns("Prod", "Cant", "Total")
        table.cursor_type = "row"
        self.refresh_confirm_table()

    def refresh_confirm_table(self):
        table = self.query_one("#pay_confirm_table")
        table.clear()
        new_total = 0.0
        for code, data in self.cart.items():
            sub = data[1] * data[2]
            new_total += sub
            table.add_row(data[0][:20], str(data[1]), f"{sub:.2f}")
        
        self.total = new_total
        self.query_one("#pay_confirm_subtotal").update(f"TOTAL A PAGAR: S/ {self.total:.2f}")
        self.query_one("#lbl_change").update(f"VUELTO: S/ {max(0, float(self.query_one('#in_paid_amount').value or 0) - self.total):.2f}")

    def on_button_pressed(self, event: Button.Pressed):
        if event.button.id == "btn_search_client":
            self.consultar_cliente()
        elif event.button.id == "btn_add_qty":
            self.modify_qty(1)
        elif event.button.id == "btn_sub_qty":
            self.modify_qty(-1)
        elif event.button.id == "btn_remove_item":
            self.remove_item()
        elif event.button.id in ["btn_boleta", "btn_factura", "btn_ticket"]:
            for bid in ["btn_boleta", "btn_factura", "btn_ticket"]: self.query_one(f"#{bid}").variant = "default"
            event.button.variant = "primary"
            self.comp_type = str(event.button.label)
        elif event.button.id in ["btn_efectivo", "btn_tarjeta", "btn_yape"]:
            for bid in ["btn_efectivo", "btn_tarjeta", "btn_yape"]: self.query_one(f"#{bid}").variant = "default"
            event.button.variant = "primary"
            self.payment_method = str(event.button.label)
        elif event.button.id == "btn_confirm_pay":
            self.finalizar_proceso()
        elif event.button.id == "btn_cancel_pay":
            self.dismiss(None)

    def finalizar_proceso(self):
        """Valida y finaliza el proceso de venta enviando los datos del cliente y pago."""
        name = self.query_one("#in_client_name").value.strip() or "PÚBLICO EN GENERAL"
        doc = self.query_one("#in_client_doc").value.strip() or "00000000"
        
        try:
            paid = float(self.query_one("#in_paid_amount").value or 0)
        except ValueError:
            self.app.notify("Monto recibido inválido.", severity="error")
            return
        
        if self.comp_type == "FACTURA" and len(doc) != 11:
            self.app.notify("RUC inválido para Factura (debe tener 11 dígitos).", severity="error")
            return

        # Guardamos solo los datos presentes en este diseño compacto
        self.dismiss({
            'metodo_pago': self.payment_method,
            'tipo_comprobante': self.comp_type,
            'monto_pagado': paid,
            'vuelto': max(0, paid - self.total),
            'cliente_nombre': name.upper(),
            'cliente_documento': doc,
            'total_final': self.total
        })

    def consultar_cliente(self):
        doc = self.query_one("#in_client_doc").value.strip()
        self.app.notify(f"Buscando {doc}...")
        c = database.buscar_cliente_local(doc)
        if c:
            self.query_one("#in_client_name").value = c['nombre']
            self.query_one("#in_client_dir").value = c.get('direccion') or ""
            self.app.notify("Cliente encontrado.")
        else:
            self.app.notify("No encontrado localmente.")

    def on_input_changed(self, event: Input.Changed):
        if event.input.id == "in_paid_amount":
            try:
                change = max(0, float(event.value or 0) - self.total)
                self.query_one("#lbl_change").update(f"VUELTO: S/ {change:.2f}")
            except ValueError: pass

    def consultar_cliente(self):
        """Consulta datos de DNI/RUC priorizando la base de datos local."""
        documento = self.query_one("#in_client_doc").value.strip()
        if not (len(documento) == 8 or len(documento) == 11):
            self.app.notify("Ingrese 8 dígitos para DNI o 11 para RUC.", severity="error")
            return

        self.app.notify("Buscando cliente...", severity="information")
        
        # 1. INTENTO LOCAL
        c = database.buscar_cliente_local(documento)
        if c:
            self.query_one("#in_client_name").value = str(c['nombre'] or "")
            self.app.notify("¡Cliente encontrado!")
            return

        # 2. INTENTO POR INTERNET
        self.app.notify("Consultando internet...", severity="information")
        tipo = "dni" if len(documento) == 8 else "ruc"
        url = f"https://api.apisperu.com/v1/{tipo}/{documento}"
        headers = {"User-Agent": "Mozilla/5.0"}

        try:
            response = requests.get(url, headers=headers, timeout=3)
            if response.status_code == 200:
                data = response.json()
                nombre = data.get('nombre') or data.get('razonSocial') or data.get('nombre_completo')
                if nombre:
                    self.query_one("#in_client_name").value = nombre.upper()
                    self.app.notify("¡Datos encontrados!")
                    return
        except Exception: pass
        
        self.app.notify("No se encontró. Ingrese manualmente.", severity="warning")
        self.query_one("#in_client_name").focus()

    def on_input_changed(self, event: Input.Changed):
        if event.input.id == "in_paid_amount":
            try:
                change = max(0, float(event.value or 0) - self.total)
                self.query_one("#lbl_change").update(f"Vuelto: S/ {change:.2f}")
            except ValueError: pass

    def modify_qty(self, delta):
        """Aumenta o disminuye la cantidad del producto seleccionado."""
        try:
            table = self.query_one("#pay_confirm_table")
            if table.cursor_row is not None:
                # El código es la clave de fila en DataTable o lo obtenemos de la base de datos (aquí usamos el nombre como referencia si no hay clave)
                # Pero lo más seguro es obtener el código del carrito basándonos en la fila
                codes = list(self.cart.keys())
                code = codes[table.cursor_row]
                
                new_qty = self.cart[code][1] + delta
                if new_qty > 0:
                    # Verificar stock
                    if new_qty <= self.cart[code][3]:
                        self.cart[code][1] = new_qty
                        self.refresh_confirm_table()
                    else:
                        self.app.notify(f"No hay más stock disponible de {self.cart[code][0]}", severity="error")
                elif new_qty == 0:
                    self.remove_item()
        except Exception as e:
            self.app.notify(f"Seleccione un producto para editar. {e}", severity="error")

    def remove_item(self):
        """Elimina el producto seleccionado del carrito de confirmación."""
        try:
            table = self.query_one("#pay_confirm_table")
            if table.cursor_row is not None:
                codes = list(self.cart.keys())
                code = codes[table.cursor_row]
                del self.cart[code]
                self.refresh_confirm_table()
                if not self.cart:
                    self.dismiss(None) # Si vacía todo, cerramos el diálogo
        except Exception:
            self.app.notify("Seleccione un producto para eliminar.", severity="error")

class FinalReceiptScreen(Screen):
    def __init__(self, cart, total, pay_data, correlativo):
        super().__init__()
        self.cart = cart
        self.total = total
        self.pay_data = pay_data
        self.correlativo = correlativo

    def compose(self) -> ComposeResult:
        receipt = self.generar_texto_ticket()
        with Vertical(classes="modal-container", id="final-receipt-container"):
            yield Label("¡VENTA FINALIZADA CON ÉXITO!", id="modal_title")
            yield Static(receipt, id="ticket_text_view")
            
            # ORGANIZACIÓN EN GRID PARA QUE TODO QUEPA BIEN
            with Grid(id="receipt-buttons-grid"):
                yield Button("Ticket 80mm", variant="success", id="btn_print_html")
                yield Button("Formato A4 (PDF)", variant="primary", id="btn_print_a4")
                yield Button("WhatsApp 📱", variant="success", id="btn_whatsapp")
                yield Button("Correo 📧", variant="primary", id="btn_email")
            
            yield Button("NUEVA VENTA / CERRAR", variant="error", id="btn_done")

    def on_button_pressed(self, event: Button.Pressed):
        if event.button.id == "btn_print_html":
            self.imprimir_ticket_html()
        elif event.button.id == "btn_print_a4":
            self.generar_pdf_a4()
        elif event.button.id == "btn_whatsapp":
            self.enviar_whatsapp()
        elif event.button.id == "btn_email":
            self.enviar_correo()
        elif event.button.id == "btn_done":
            self.app.pop_screen()

    def generar_pdf_a4(self):
        """Genera un PDF profesional en formato A4."""
        try:
            os.makedirs("ventas", exist_ok=True)
            filename = os.path.join("ventas", f"Factura_A4_{self.correlativo}.pdf")
            
            pdf = FPDF()
            pdf.add_page()
            
            # Encabezado Empresa
            pdf.set_font("Arial", "B", 16)
            pdf.cell(0, 10, "FERRETERÍA ROLIK", ln=True, align="L")
            pdf.set_font("Arial", "", 10)
            pdf.cell(0, 5, "RUC: 10440809320", ln=True)
            pdf.cell(0, 5, "Dirección: Mz A lt 26 P.I. Madera KM 15.5", ln=True)
            pdf.cell(0, 5, "Tel: 988352912 / 932326764", ln=True)
            
            # Recuadro del Comprobante (Derecha)
            pdf.set_xy(130, 10)
            pdf.set_font("Arial", "B", 12)
            pdf.cell(70, 25, "", border=1) # Recuadro
            pdf.set_xy(130, 12)
            pdf.cell(70, 7, f"R.U.C. 10440809320", ln=True, align="C")
            pdf.set_xy(130, 19)
            pdf.cell(70, 7, self.pay_data['tipo_comprobante'], ln=True, align="C")
            pdf.set_xy(130, 26)
            pdf.cell(70, 7, self.correlativo, ln=True, align="C")
            
            pdf.set_xy(10, 45)
            pdf.set_font("Arial", "B", 10)
            pdf.cell(0, 10, "DATOS DEL CLIENTE", ln=True)
            pdf.set_font("Arial", "", 10)
            pdf.cell(0, 6, f"Señor(es): {self.pay_data.get('cliente_nombre', 'PÚBLICO GENERAL')}", ln=True)
            pdf.cell(0, 6, f"Documento: {self.pay_data.get('cliente_documento', '00000000')}", ln=True)
            pdf.cell(0, 6, f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True)
            pdf.ln(5)

            # Tabla de Productos
            pdf.set_font("Arial", "B", 10)
            pdf.set_fill_color(240, 240, 240)
            pdf.cell(20, 10, "Cant.", 1, 0, "C", True)
            pdf.cell(110, 10, "Descripción", 1, 0, "L", True)
            pdf.cell(30, 10, "P. Unit", 1, 0, "R", True)
            pdf.cell(30, 10, "Total", 1, 1, "R", True)

            pdf.set_font("Arial", "", 10)
            for _, item in self.cart.items():
                # [nombre, qty, price, stock, marca, unidad]
                sub = item[1] * item[2]
                pdf.cell(20, 8, str(item[1]), 1, 0, "C")
                pdf.cell(110, 8, str(item[0])[:50], 1, 0, "L")
                pdf.cell(30, 8, f"{item[2]:.2f}", 1, 0, "R")
                pdf.cell(30, 8, f"{sub:.2f}", 1, 1, "R")

            # Totales
            pdf.ln(5)
            subtotal = self.total / 1.18
            igv = self.total - subtotal
            pdf.set_font("Arial", "B", 10)
            pdf.cell(160, 8, "SUBTOTAL S/", 0, 0, "R")
            pdf.cell(30, 8, f"{subtotal:.2f}", 1, 1, "R")
            pdf.cell(160, 8, "I.G.V. (18%) S/", 0, 0, "R")
            pdf.cell(30, 8, f"{igv:.2f}", 1, 1, "R")
            pdf.cell(160, 8, "TOTAL FINAL S/", 0, 0, "R")
            pdf.set_fill_color(255, 255, 200)
            pdf.cell(30, 8, f"{self.total:.2f}", 1, 1, "R", True)

            pdf.output(filename)
            self.app.notify(f"PDF A4 generado: {filename}", severity="information")
            if os.name == 'nt': os.startfile(filename)
        except Exception as e:
            self.app.notify(f"Error PDF A4: {e}", severity="error")

    def enviar_whatsapp(self):
        """Abre WhatsApp Web con un mensaje predeterminado."""
        telefono = self.pay_data.get('cliente_tel', '').strip()
        if not telefono or len(telefono) < 9:
            self.app.notify("El cliente no tiene un teléfono válido registrado.", severity="warning")
            return
        
        mensaje = f"Hola *{self.pay_data['cliente_nombre']}*, le enviamos su comprobante *{self.correlativo}* de *FERRETERÍA ROLIK* por un total de *S/ {self.total:.2f}*. ¡Gracias por su compra!"
        import urllib.parse
        encoded_msg = urllib.parse.quote(mensaje)
        url = f"https://web.whatsapp.com/send?phone=51{telefono}&text={encoded_msg}"
        webbrowser.open(url)
        self.app.notify("Abriendo WhatsApp Web...")

    def enviar_correo(self):
        """Prepara un correo electrónico con los datos de la venta."""
        email = self.pay_data.get('cliente_email', '').strip()
        if not email:
            self.app.notify("El cliente no tiene un correo registrado.", severity="warning")
            return
        
        subject = f"Comprobante de Pago {self.correlativo} - FERRETERIA ROLIK"
        body = f"Estimado cliente,\n\nSe adjunta la informacion de su compra por S/ {self.total:.2f}.\nNumero de comprobante: {self.correlativo}\n\nGracias por su preferencia."
        import urllib.parse
        url = f"mailto:{email}?subject={urllib.parse.quote(subject)}&body={urllib.parse.quote(body)}"
        webbrowser.open(url)
        self.app.notify("Abriendo gestor de correo...")

    def generar_html_ticket(self):
        """Genera un archivo HTML con estilo profesional para ticketeras de 80mm."""
        now = datetime.now().strftime('%d/%m/%Y %H:%M')
        items_html = ""
        for _, datos in self.cart.items():
            # [nombre, qty, price, stock, marca, unidad]
            nombre = str(datos[0])[:25]
            qty = datos[1]
            price = datos[2]
            sub = qty * price
            items_html += f"""
            <tr>
                <td>{qty}</td>
                <td>{nombre}</td>
                <td style='text-align: right;'>{sub:.2f}</td>
            </tr>"""

        subtotal = self.total / 1.18
        igv = self.total - subtotal

        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                @page {{ margin: 0; }}
                body {{ 
                    width: 72mm; 
                    font-family: 'Arial Narrow', Arial, sans-serif; 
                    font-size: 12px; 
                    margin: 0; 
                    padding: 4mm;
                    color: black;
                }}
                .center {{ text-align: center; }}
                .bold {{ font-weight: bold; }}
                .line {{ border-top: 1px dashed black; margin: 5px 0; }}
                table {{ width: 100%; border-collapse: collapse; margin: 10px 0; }}
                th {{ border-bottom: 1px solid black; text-align: left; }}
                .total-table {{ margin-top: 10px; float: right; width: 60%; }}
                .footer {{ margin-top: 20px; font-size: 10px; }}
            </style>
        </head>
        <body onload="window.print();">
            <div class="center">
                <div class="bold" style="font-size: 16px;">FERRETERÍA ROLIK</div>
                <div>RUC: 10440809320</div>
                <div>Mz A lt 26 P.I. Madera KM 15.5</div>
                <div>Cel: 988352912 / 932326764</div>
                <div class="line"></div>
                <div class="bold">{self.pay_data['tipo_comprobante']}</div>
                <div>{self.correlativo}</div>
                <div>Fecha: {now}</div>
            </div>
            
            <div style="margin-top: 10px;">
                <div><b>Cliente:</b> {self.pay_data.get('cliente_nombre', 'PÚBLICO GENERAL')}</div>
                <div><b>Doc:</b> {self.pay_data.get('cliente_documento', '00000000')}</div>
            </div>

            <table>
                <thead>
                    <tr>
                        <th style="width: 15%;">Cant</th>
                        <th style="width: 60%;">Descrip</th>
                        <th style="width: 25%; text-align: right;">Total</th>
                    </tr>
                </thead>
                <tbody>
                    {items_html}
                </tbody>
            </table>

            <div class="line"></div>
            <table class="total-table">
                <tr><td>SUBTOTAL:</td><td style="text-align: right;">S/ {subtotal:.2f}</td></tr>
                <tr><td>IGV (18%):</td><td style="text-align: right;">S/ {igv:.2f}</td></tr>
                <tr class="bold" style="font-size: 14px;"><td>TOTAL:</td><td style="text-align: right;">S/ {self.total:.2f}</td></tr>
            </table>
            
            <div style="clear: both; margin-top: 10px;">
                <div><b>Pago:</b> {self.pay_data['metodo_pago']}</div>
                <div><b>Recibido:</b> S/ {self.pay_data['monto_pagado']:.2f}</div>
                <div><b>Vuelto:</b> S/ {self.pay_data['vuelto']:.2f}</div>
            </div>

            <div class="center footer">
                <div class="line"></div>
                <p>Representación impresa de la {self.pay_data['tipo_comprobante'].lower()}</p>
            </div>
        </body>
        </html>
        """
        return html

    def on_button_pressed(self, event: Button.Pressed):
        if event.button.id == "btn_print_thermal":
            self.imprimir_ticket_archivo()
        elif event.button.id == "btn_print_html":
            self.imprimir_ticket_html()
        else:
            self.app.pop_screen()

    def generar_texto_ticket(self):
        """Genera el contenido del ticket compacto para ticketera de 80mm."""
        w = 36 # Ancho reducido para evitar desbordes
        now = datetime.now().strftime('%d/%m/%y %H:%M')
        
        t = f"{'='*w}\n"
        t += f"{'FERRETERIA ROLIK':^{w}}\n"
        t += f"{'RUC: 10440809320':^{w}}\n"
        t += f"{'CEL: 988352912 / 932326764':^{w}}\n"
        t += f"{'MZ A LT 26 P.I. MADERA KM 15.5':^{w}}\n"
        t += f"{'='*w}\n"
        t += f"COMP: {self.pay_data['tipo_comprobante']}\n"
        t += f"NUM : {self.correlativo}\n"
        t += f"FECHA: {now}\n"
        t += f"CLI : {self.pay_data.get('cliente_nombre', 'P. GENERAL')[:w-6]}\n"
        t += f"DOC : {self.pay_data.get('cliente_documento', '00000000')}\n"
        t += f"{'-'*w}\n"
        t += f"{'CANT':<4} {'DESCRIPCION':<22} {'TOT':>8}\n"
        t += f"{'-'*w}\n"
        
        for codigo, datos in self.cart.items():
            # [nombre, qty, price, stock, marca, unidad]
            nombre = str(datos[0])[:20]
            qty = datos[1]
            price = datos[2]
            sub = qty * price
            t += f"{str(qty):<4} {nombre:<22} {sub:>8.2f}\n"
        
        t += f"{'-'*w}\n"
        subtotal = self.total / 1.18
        igv = self.total - subtotal
        
        t += f"{'SUBTOTAL:':<26} {subtotal:>8.2f}\n"
        t += f"{'IGV (18%):':<26} {igv:>8.2f}\n"
        t += f"{'TOTAL:':<24} S/ {self.total:>8.2f}\n"
        t += f"{'-'*w}\n"
        t += f"PAGO: {self.pay_data['metodo_pago']} | REC: {self.pay_data['monto_pagado']:.2f}\n"
        t += f"VUELTO: S/ {self.pay_data['vuelto']:.2f}\n"
        t += f"{'='*w}\n\n\n" # Espacio final para el corte manual
        return t

    def imprimir_ticket_archivo(self):
        """Guarda el ticket en un archivo .txt listo para enviar a la ticketera."""
        try:
            os.makedirs("ventas", exist_ok=True)
            filename = os.path.join("ventas", f"Ticket_{self.correlativo}.txt")
            content = self.generar_texto_ticket()
            with open(filename, "w", encoding="utf-8") as f:
                f.write(content)
            
            self.app.notify(f"Ticket generado: {filename}", severity="information")
            if os.name == 'nt':
                os.startfile(filename)
        except Exception as e:
            self.app.notify(f"Error al imprimir: {e}", severity="error")

    def imprimir_ticket_html(self):
        """Genera y abre el ticket en formato HTML."""
        try:
            os.makedirs("ventas", exist_ok=True)
            filename = os.path.join("ventas", f"Ticket_{self.correlativo}.html")
            content = self.generar_html_ticket()
            with open(filename, "w", encoding="utf-8") as f:
                f.write(content)
            
            self.app.notify(f"HTML generado: {filename}", severity="information")
            if os.name == 'nt':
                os.startfile(filename)
        except Exception as e:
            self.app.notify(f"Error HTML: {e}", severity="error")

class CashScreen(Screen):
    BINDINGS = [Binding("escape", "app.pop_screen", "Volver")]
    
    def compose(self) -> ComposeResult:
        with Vertical(classes="modal-container", id="cash-panel"):
            yield Label("GESTIÓN PROFESIONAL DE CAJA", id="modal_title")
            yield Static(id="cash_status_label", classes="status-banner")
            
            # Panel de Apertura
            with Vertical(id="open_cash_container"):
                yield Label("Monto Inicial (Fondo de Sencillo):")
                yield Input("0.00", id="in_initial_fund", restrict=r"[0-9.]*")
                yield Button("ABRIR CAJA", variant="success", id="btn_open_cash")
            
            # Panel de Operaciones (Solo si está abierta)
            with Vertical(id="ops_cash_container"):
                with Horizontal(classes="cash-ops-row"):
                    yield Button("INGRESAR DINERO", id="btn_cash_in", variant="primary")
                    yield Button("RETIRAR DINERO", id="btn_cash_out", variant="error")
                
                with Horizontal(classes="cash-ops-row"):
                    yield Button("CORTE X (Ver Resumen)", id="btn_corte_x", variant="warning")
                    yield Button("CERRAR CAJA (Corte Z)", id="btn_close_cash", variant="error")
            
            yield Button("HISTORIAL DE CAJAS", id="btn_cash_history", variant="primary")
            yield Button("VOLVER (ESC)", variant="error", id="btn_back")
    
    def on_mount(self): 
        self.refresh_cash_status()
    
    def refresh_cash_status(self):
        session = database.get_active_session()
        is_open = bool(session)
        self.query_one("#open_cash_container").display = not is_open
        self.query_one("#ops_cash_container").display = is_open
        
        banner = self.query_one("#cash_status_label")
        if is_open:
            banner.update(f"ESTADO: CAJA ABIERTA | Inicio: {session['open_date'][:16]}")
            banner.add_class("status-open")
        else:
            banner.update("ESTADO: CAJA CERRADA")
            banner.remove_class("status-open")
    
    def on_button_pressed(self, event: Button.Pressed):
        session = database.get_active_session()
        
        if event.button.id == "btn_open_cash":
            fund = float(self.query_one("#in_initial_fund").value or 0)
            database.open_cash_session(fund, self.app.user_id)
            self.app.notify("Caja abierta correctamente.", severity="success")
            self.refresh_cash_status()
            
        elif event.button.id == "btn_cash_in":
            self.app.push_screen(MovementDialog("INGRESO"), lambda _: self.refresh_cash_status())
            
        elif event.button.id == "btn_cash_out":
            self.app.push_screen(MovementDialog("RETIRO"), lambda _: self.refresh_cash_status())
            
        elif event.button.id == "btn_corte_x":
            if session: self.app.push_screen(CorteXDialog(session['id']))
            
        elif event.button.id == "btn_close_cash":
            if session:
                summary = database.get_cash_session_summary(session['id'])
                database.close_cash_session(session['id'], summary['total_general'], self.app.user_id)
                self.app.notify("Caja cerrada. Se ha generado el reporte final.", severity="success")
                self.refresh_cash_status()
        
        elif event.button.id == "btn_cash_history":
            self.app.push_screen(CashReportScreen())
            
        elif event.button.id == "btn_back":
            self.app.pop_screen()

class MovementDialog(Screen):
    """Diálogo para registrar ingresos o retiros manuales."""
    def __init__(self, tipo):
        super().__init__()
        self.tipo = tipo

    def compose(self) -> ComposeResult:
        with Vertical(classes="modal-container", id="mov-dialog"):
            yield Label(f"REGISTRAR {self.tipo}", id="modal_title")
            yield Label("Monto:")
            yield Input("0.00", id="mov_amount", restrict=r"[0-9.]*")
            yield Label("Descripción / Motivo:")
            yield Input(placeholder="Ej: Pago de luz, Sencillo, etc.", id="mov_desc")
            with Horizontal(classes="form-buttons"):
                yield Button("Guardar", variant="success", id="btn_save_mov")
                yield Button("Cancelar", variant="error", id="btn_cancel_mov")

    def on_button_pressed(self, event: Button.Pressed):
        if event.button.id == "btn_save_mov":
            monto = float(self.query_one("#mov_amount").value or 0)
            desc = self.query_one("#mov_desc").value.strip()
            session = database.get_active_session()
            if monto > 0 and session:
                database.add_cash_movement(session['id'], self.app.user_id, self.tipo, monto, desc)
                self.app.notify(f"{self.tipo} registrado correctamente.")
                self.dismiss(True)
        else:
            self.dismiss(False)

class CorteXDialog(Screen):
    """Diálogo que muestra el resumen en tiempo real de la caja."""
    def __init__(self, session_id):
        super().__init__()
        self.session_id = session_id

    def compose(self) -> ComposeResult:
        with Vertical(classes="modal-container", id="cortex-dialog"):
            yield Label("CORTE X - RESUMEN DE CAJA", id="modal_title")
            yield Static(id="cortex_content", classes="receipt-text-area")
            yield Button("Cerrar", variant="primary", id="btn_close_cortex")

    def on_mount(self):
        res = database.get_cash_session_summary(self.session_id)
        
        text = f"{'='*36}\n"
        text += f"FONDO INICIAL:   S/ {res['inicial']:>10.2f}\n"
        text += f"{'-'*36}\n"
        text += "VENTAS POR METODO:\n"
        for metodo, total in res['ventas'].items():
            text += f"  {metodo:<15} S/ {total:>10.2f}\n"
        
        text += f"{'-'*36}\n"
        text += "MOVIMIENTOS MANUALES:\n"
        text += f"  INGRESOS (+)    S/ {res['movimientos'].get('INGRESO', 0):>10.2f}\n"
        text += f"  RETIROS  (-)    S/ {res['movimientos'].get('RETIRO', 0):>10.2f}\n"
        
        text += f"{'='*36}\n"
        text += f"EFECTIVO ESPERADO: S/ {res['efectivo_esperado']:>10.2f}\n"
        text += f"TOTAL GENERAL:     S/ {res['total_general']:>10.2f}\n"
        text += f"{'='*36}\n"
        
        self.query_one("#cortex_content").update(text)

    def on_button_pressed(self, event: Button.Pressed):
        self.dismiss()

# --- Módulo de Órdenes de Compra ---
class SalesReportScreen(Screen):
    BINDINGS = [Binding("escape", "app.pop_screen", "Volver"), ("r", "refresh_report", "Refrescar")]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.selected_user_id = None
        self.start_date = ""
        self.end_date = ""

    def compose(self) -> ComposeResult:
        yield Header()
        yield Label("REPORTE DE VENTAS POR VENDEDOR", id="section_title")
        with Horizontal(id="report_filters"):
            yield Label("Vendedor:")
            yield Input(placeholder="Todos", id="filter_seller_name", classes="filter-input")
            yield Button("Seleccionar Vendedor", id="btn_select_seller")
            yield Label("Desde (YYYY-MM-DD):")
            yield Input(placeholder="YYYY-MM-DD", id="filter_start_date", classes="filter-input")
            yield Label("Hasta (YYYY-MM-DD):")
            yield Input(placeholder="YYYY-MM-DD", id="filter_end_date", classes="filter-input")
            yield Button("Refrescar (R)", id="btn_refresh_report")
        yield DataTable(id="sales_report_table")
        yield Footer()

    def on_mount(self) -> None:
        table = self.query_one(DataTable)
        table.cursor_type = "row"
        table.add_columns("ID Venta", "Fecha", "Total", "Vendedor")
        self.refresh_report()

    def refresh_report(self):
        try:
            start_date = self.query_one("#filter_start_date", Input).value
            end_date = self.query_one("#filter_end_date", Input).value
            
            table = self.query_one(DataTable)
            table.clear()

            sales = database.get_sales_history(
                user_id=self.selected_user_id,
                start_date=start_date if start_date else None,
                end_date=end_date if end_date else None
            )

            total_sales_amount = 0.0
            for sale in sales:
                table.add_row(
                    str(sale['transaction_id']),
                    str(sale['transaction_date']),
                    f"${sale['transaction_total']:.2f}",
                    str(sale['seller_name'])
                )
                total_sales_amount += sale['transaction_total']
            
            table.add_row("", "", "", "", key="total-row")
            table.add_row("TOTAL:", "", "", f"${total_sales_amount:.2f}", key="total-row-val")

        except Exception as e:
            self.app.notify(f"Error al cargar el reporte de ventas: {e}", severity="error")

    def action_refresh_report(self):
        self.refresh_report()

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "btn_refresh_report":
            self.refresh_report()
        elif event.button.id == "btn_select_seller":
            self.app.push_screen(UserSelectionDialog(), callback=self.on_seller_selected)
        elif event.button.id == "btn_back":
            self.app.pop_screen()

    def on_seller_selected(self, user_data: dict | None):
        if user_data:
            self.selected_user_id = user_data['id']
            self.query_one("#filter_seller_name", Input).value = user_data['username']
        else:
            self.selected_user_id = None
            self.query_one("#filter_seller_name", Input).value = "Todos"
        self.refresh_report()

class CashReportScreen(Screen):
    BINDINGS = [Binding("escape", "app.pop_screen", "Volver"), ("r", "refresh_report", "Refrescar")]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.selected_user_id = None
        self.start_date = ""
        self.end_date = ""

    def compose(self) -> ComposeResult:
        yield Header()
        yield Label("REPORTE DE SESIONES DE CAJA", id="section_title")
        with Horizontal(id="report_filters"):
            yield Label("Usuario:")
            yield Input(placeholder="Todos", id="filter_user_name", classes="filter-input")
            yield Button("Seleccionar Usuario", id="btn_select_user")
            yield Label("Desde (YYYY-MM-DD):")
            yield Input(placeholder="YYYY-MM-DD", id="filter_start_date", classes="filter-input")
            yield Label("Hasta (YYYY-MM-DD):")
            yield Input(placeholder="YYYY-MM-DD", id="filter_end_date", classes="filter-input")
            yield Button("Refrescar (R)", id="btn_refresh_report")
        yield DataTable(id="cash_report_table")
        yield Footer()

    def on_mount(self) -> None:
        table = self.query_one(DataTable)
        table.cursor_type = "row"
        table.add_columns("ID Sesión", "Apertura", "Cierre", "Monto Inicial", "Ventas Totales", "Estado", "Abrió", "Cerró")
        self.refresh_report()

    def refresh_report(self):
        try:
            start_date = self.query_one("#filter_start_date", Input).value
            end_date = self.query_one("#filter_end_date", Input).value
            
            table = self.query_one(DataTable)
            table.clear()

            sessions = database.get_cash_sessions_history(
                user_id=self.selected_user_id,
                start_date=start_date if start_date else None,
                end_date=end_date if end_date else None
            )

            for session in sessions:
                table.add_row(
                    str(session['session_id']),
                    str(session['open_date']),
                    str(session['close_date'] or "N/A"),
                    f"${session['initial_fund']:.2f}",
                    f"${session['total_sales']:.2f}",
                    str(session['status']),
                    str(session['opened_by_username'] or "N/A"),
                    str(session['closed_by_username'] or "N/A")
                )
        except Exception as e:
            self.app.notify(f"Error al cargar el reporte de caja: {e}", severity="error")

    def action_refresh_report(self):
        self.refresh_report()

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "btn_refresh_report":
            self.refresh_report()
        elif event.button.id == "btn_select_user":
            self.app.push_screen(UserSelectionDialog(), callback=self.on_user_selected)
        elif event.button.id == "btn_back":
            self.app.pop_screen()

    def on_user_selected(self, user_data: dict | None):
        if user_data:
            self.selected_user_id = user_data['id']
            self.query_one("#filter_user_name", Input).value = user_data['username']
        else:
            self.selected_user_id = None
            self.query_one("#filter_user_name", Input).value = "Todos"
        self.refresh_report()

class UserSelectionDialog(Screen):
    """Diálogo para seleccionar un usuario de la lista."""
    def compose(self) -> ComposeResult:
        yield Label("SELECCIONAR USUARIO", id="modal_title")
        yield Input(placeholder="Filtrar por nombre de usuario...", id="filter_user_input")
        yield DataTable(id="user_selection_table")
        with Horizontal(classes="form-buttons"):
            yield Button("Seleccionar", "success", id="btn_select_user_dialog")
            yield Button("Cancelar", "error", id="btn_cancel_user_dialog")
            yield Button("Limpiar Selección", "warning", id="btn_clear_user_selection")

    def on_mount(self) -> None:
        self.query_one("#user_selection_table").add_columns("ID", "Usuario", "Rol")
        self.refresh_user_list()
        self.query_one("#user_selection_table").cursor_type = "row"

    def refresh_user_list(self, filter_text: str = ""):
        table = self.query_one("#user_selection_table")
        table.clear()
        users = database.get_all_users()
        for u in users:
            if filter_text.lower() in u['username'].lower():
                table.add_row(str(u['id']), u['username'], u['role'], key=str(u['id']))
        table.focus() # Enfocar la tabla para navegación con flechas

    def on_input_changed(self, event: Input.Changed) -> None:
        if event.input.id == "filter_user_input":
            self.refresh_user_list(event.value)

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "btn_select_user_dialog":
            table = self.query_one("#user_selection_table")
            if table.row_count > 0 and table.cursor_row < table.row_count:
                user_id = int(table.get_cell_at((table.cursor_row, 0)))
                user = database.get_user_by_id(user_id)
                self.dismiss(user)
            else:
                self.app.notify("Selecciona un usuario o cancela.", severity="warning")
        elif event.button.id == "btn_clear_user_selection":
            self.dismiss(None) # Devuelve None para indicar que no hay selección
        elif event.button.id == "btn_cancel_user_dialog":
            self.dismiss(None)

class ReportsMenuScreen(Screen):
    BINDINGS = [Binding("escape", "app.pop_screen", "Volver")]

    def compose(self) -> ComposeResult:
        yield Header()
        yield Label("ROLIK - CENTRO DE REPORTES", id="section_title")
        
        with VerticalScroll(id="report_menu_container"):
            with Horizontal(classes="report-row"):
                yield Button("Ventas de Hoy", variant="primary", id="btn_rep_today")
                yield Button("Ventas por Rango / Ganancias", variant="primary", id="btn_rep_range")
            
            with Horizontal(classes="report-row"):
                yield Button("Ventas por Producto", variant="primary", id="btn_rep_prod")
                yield Button("Top 10 Más Vendidos", variant="primary", id="btn_rep_top")
            
            with Horizontal(classes="report-row"):
                yield Button("Bajo Stock / Reabastecer", variant="warning", id="btn_rep_stock")
                yield Button("Kardex (Movimientos)", variant="warning", id="btn_rep_kardex")
            
            with Horizontal(classes="report-row"):
                yield Button("Historial de Recibos", variant="success", id="btn_rep_receipts")
                yield Button("Ventas por Vendedor", variant="success", id="btn_rep_sellers")
            
            with Horizontal(classes="report-row"):
                yield Button("Volver al Menú", variant="error", id="btn_back")
        yield Footer()

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "btn_rep_today":
            self.app.push_screen(DailySalesReportScreen())
        elif event.button.id == "btn_rep_range":
            self.app.push_screen(FinancialReportScreen())
        elif event.button.id == "btn_rep_prod":
            self.app.push_screen(ProductSalesReportScreen())
        elif event.button.id == "btn_rep_stock":
            self.app.push_screen(LowStockReportScreen())
        elif event.button.id == "btn_rep_kardex":
            self.app.push_screen(KardexReportScreen())
        elif event.button.id == "btn_rep_top":
            self.app.push_screen(TopProductsReportScreen())
        elif event.button.id == "btn_rep_receipts":
            self.app.push_screen(SalesHistoryReportScreen())
        elif event.button.id == "btn_rep_sellers":
            self.app.push_screen(SalesReportScreen())
        elif event.button.id == "btn_back":
            self.app.pop_screen()

# --- PANTALLAS DE REPORTES ESPECÍFICOS ---

class ProductSalesReportScreen(Screen):
    def compose(self) -> ComposeResult:
        yield Header()
        yield Label("REPORTE DE VENTAS POR PRODUCTO", id="section_title")
        with Horizontal(id="filter_bar"):
            yield Input(placeholder="Inicio YYYY-MM-DD", id="rep_start")
            yield Input(placeholder="Fin YYYY-MM-DD", id="rep_end")
            yield Button("Filtrar", variant="success", id="btn_filter")
        yield DataTable(id="product_sales_table")
        yield Button("Cerrar", variant="error", id="btn_close")
        yield Footer()

    def on_mount(self):
        table = self.query_one(DataTable)
        table.add_columns("Producto", "Cant. Vendida", "Total Generado", "Margen Ganancia")
        table.cursor_type = "row"

    def on_button_pressed(self, event: Button.Pressed):
        if event.button.id == "btn_filter":
            start = self.query_one("#rep_start").value
            end = self.query_one("#rep_end").value
            table = self.query_one(DataTable)
            table.clear()
            for p in database.get_report_sales_by_product(start, end):
                table.add_row(
                    p['nombre'], 
                    str(p['cant_vendida']), 
                    f"S/ {p['total_generado']:.2f}", 
                    f"S/ {p['margen_ganancia']:.2f}"
                )
        else:
            self.app.pop_screen()

class DailySalesReportScreen(Screen):
    def compose(self) -> ComposeResult:
        yield Header()
        yield Label("RESUMEN DE VENTAS DEL DÍA", id="section_title")
        yield Static(id="daily_stats", classes="report-summary-box")
        yield Button("Cerrar", variant="error", id="btn_close")
        yield Footer()

    def on_mount(self):
        data = database.get_report_daily_sales()
        if data:
            stats = f"""
TOTAL TRANSACCIONES: {data['total_transacciones']}
-------------------------------------------
MONTO TOTAL VENDIDO: S/ {data['monto_total'] or 0.0:.2f}
-------------------------------------------
PAGOS EN EFECTIVO:   S/ {data['efectivo'] or 0.0:.2f}
PAGOS EN TARJETA:    S/ {data['tarjeta'] or 0.0:.2f}
OTROS MEDIOS:        S/ {data['otros'] or 0.0:.2f}
"""
            self.query_one("#daily_stats").update(stats)

    def on_button_pressed(self, event: Button.Pressed):
        self.app.pop_screen()

class FinancialReportScreen(Screen):
    def compose(self) -> ComposeResult:
        yield Header()
        yield Label("REPORTE FINANCIERO Y GANANCIAS", id="section_title")
        with Horizontal(id="filter_bar"):
            yield Input(placeholder="Inicio YYYY-MM-DD", id="rep_start")
            yield Input(placeholder="Fin YYYY-MM-DD", id="rep_end")
            yield Button("Calcular", variant="success", id="btn_calc")
        yield Static(id="finance_results", classes="report-summary-box")
        yield Button("Cerrar", variant="error", id="btn_close")
        yield Footer()

    def on_button_pressed(self, event: Button.Pressed):
        if event.button.id == "btn_calc":
            start = self.query_one("#rep_start").value
            end = self.query_one("#rep_end").value
            data = database.get_report_sales_by_range(start, end)
            if data:
                res = f"""
INGRESOS BRUTOS:     S/ {data['ingresos_brutos'] or 0.0:.2f}
TOTAL IGV (18%):     S/ {data['total_igv'] or 0.0:.2f}
TOTAL NETO:          S/ {data['total_neto'] or 0.0:.2f}
-------------------------------------------
GANANCIA ESTIMADA:   S/ {data['ganancia_estimada'] or 0.0:.2f}
(Venta - Costo Compra)
"""
                self.query_one("#finance_results").update(res)
        else:
            self.app.pop_screen()

class LowStockReportScreen(Screen):
    def compose(self) -> ComposeResult:
        yield Header()
        yield Label("PRODUCTOS CON BAJO STOCK (ALERTA)", id="section_title")
        yield DataTable(id="low_stock_table")
        yield Button("Cerrar", variant="error", id="btn_close")
        yield Footer()

    def on_mount(self):
        table = self.query_one(DataTable)
        table.add_columns("Código", "Producto", "Stock Actual", "Mínimo", "Faltante")
        for p in database.get_report_low_stock():
            table.add_row(p['codigo'], p['nombre'], str(p['stock']), str(p['stock_minimo']), str(p['faltante']))

    def on_button_pressed(self, event: Button.Pressed):
        self.app.pop_screen()

class KardexReportScreen(Screen):
    def compose(self) -> ComposeResult:
        yield Header()
        yield Label("KARDEX - HISTORIAL DE MOVIMIENTOS", id="section_title")
        yield Label("(Presione ENTER sobre una VENTA para ver su recibo)", classes="help-text")
        yield DataTable(id="kardex_table")
        yield Button("Cerrar", variant="error", id="btn_close")
        yield Footer()

    def on_mount(self):
        table = self.query_one(DataTable)
        table.add_columns("Fecha", "Tipo", "Producto", "Cant.", "Precio Ref.", "Ref_ID")
        table.cursor_type = "row"
        for m in database.get_report_kardex():
            table.add_row(
                str(m['date'])[:16], 
                m['tipo'], 
                m['producto_codigo'], 
                str(m['cant']), 
                f"S/ {m['precio']:.2f}",
                str(m['ref_id']),
                key=f"{m['tipo']}_{m['ref_id']}"
            )

    def on_data_table_row_selected(self, event: DataTable.RowSelected):
        key_parts = str(event.row_key.value).split("_")
        tipo = key_parts[0]
        ref_id = int(key_parts[1])
        if "VENTA" in tipo:
            self.app.push_screen(ViewReceiptDialog(ref_id))
        else:
            self.app.notify("Este movimiento es una compra (OC).", severity="information")

    def on_button_pressed(self, event: Button.Pressed):
        self.app.pop_screen()

class TopProductsReportScreen(Screen):
    def compose(self) -> ComposeResult:
        yield Header()
        yield Label("TOP 10 PRODUCTOS MÁS VENDIDOS", id="section_title")
        yield DataTable(id="top_table")
        yield Button("Cerrar", variant="error", id="btn_close")
        yield Footer()

    def on_mount(self):
        table = self.query_one(DataTable)
        table.add_columns("Ranking", "Producto", "Cantidad Vendida")
        for i, p in enumerate(database.get_report_top_products(), 1):
            table.add_row(str(i), p['nombre'], str(p['total_qty']))

    def on_button_pressed(self, event: Button.Pressed):
        self.app.pop_screen()

class VoidReasonDialog(Screen):
    """Diálogo para ingresar el motivo de anulación de una venta."""
    def compose(self) -> ComposeResult:
        with Vertical(classes="modal-container", id="void-reason-container"):
            yield Label("ANULACIÓN DE VENTA", id="modal_title")
            yield Label("¿Por qué desea anular esta venta?")
            yield Input(placeholder="Ingrese el motivo aquí...", id="in_void_reason")
            with Horizontal(classes="form-buttons"):
                yield Button("ANULAR VENTA", variant="error", id="btn_confirm_void")
                yield Button("Cancelar", variant="primary", id="btn_cancel_void")

    def on_button_pressed(self, event: Button.Pressed):
        if event.button.id == "btn_confirm_void":
            reason = self.query_one("#in_void_reason").value.strip()
            if len(reason) < 5:
                self.app.notify("Por favor, ingrese un motivo válido (mín. 5 letras).", severity="error")
                return
            self.dismiss(reason)
        else:
            self.dismiss(None)

class SalesHistoryReportScreen(Screen):
    """Muestra la lista de todos los comprobantes de venta generados con opción a anulación."""
    def compose(self) -> ComposeResult:
        yield Header()
        yield Label("HISTORIAL DE COMPROBANTES GENERADOS", id="section_title")
        yield DataTable(id="sales_history_table")
        with Horizontal(classes="form-buttons"):
            yield Button("Ver Recibo (V)", variant="primary", id="btn_view_receipt")
            yield Button("Editar Recibo (E)", variant="warning", id="btn_edit_receipt")
            yield Button("ANULAR VENTA", variant="error", id="btn_void_sale")
            yield Button("Cerrar", variant="error", id="btn_close")
        yield Footer()

    def on_mount(self):
        table = self.query_one(DataTable)
        # 6 columnas: 0:Fecha, 1:Comp, 2:Cliente, 3:Total, 4:Estado, 5:ID
        table.add_columns("Fecha", "Comprobante", "Cliente", "Total", "Estado", "ID")
        table.cursor_type = "row"
        self.refresh_history()

    def refresh_history(self):
        table = self.query_one(DataTable)
        table.clear()
        
        conn = database.get_connection()
        sales = conn.execute("""
            SELECT t.id, t.date, t.total, t.status, t.tipo_comprobante, t.correlativo, t.cliente_nombre
            FROM transactions t
            ORDER BY t.date DESC
        """).fetchall()
        conn.close()

        for s in sales:
            status_text = "COMPLETADA" if s['status'] != 'VOIDED' else "[ANULADA]"
            comp = f"{s['tipo_comprobante']} {s['correlativo']}"
            cliente = str(s['cliente_nombre'] or "P. GENERAL")[:20]
            
            table.add_row(
                str(s['date'])[:16],
                comp,
                cliente,
                f"S/ {s['total']:.2f}",
                status_text,
                str(s['id']),
                key=str(s['id'])
            )

    def on_button_pressed(self, event: Button.Pressed):
        if event.button.id == "btn_view_receipt":
            self.action_view_receipt()
        elif event.button.id == "btn_edit_receipt":
            self.action_edit_receipt()
        elif event.button.id == "btn_void_sale":
            self.action_void_sale()
        elif event.button.id == "btn_close":
            self.app.pop_screen()

    def action_view_receipt(self):
        try:
            table = self.query_one(DataTable)
            if table.cursor_row is not None:
                trans_id = int(table.get_cell_at((table.cursor_row, 5)))
                self.app.push_screen(ViewReceiptDialog(trans_id))
        except Exception:
            self.app.notify("Seleccione una venta para visualizar.", severity="error")

    def action_edit_receipt(self):
        try:
            table = self.query_one(DataTable)
            if table.cursor_row is not None:
                trans_id = int(table.get_cell_at((table.cursor_row, 5)))
                self.app.push_screen(EditSaleScreen(trans_id), lambda success: self.refresh_history() if success else None)
        except Exception:
            self.app.notify("Seleccione una venta para editar.", severity="error")

    def action_void_sale(self):
        if self.app.user_role != 'admin':
            self.app.notify("Solo un administrador puede anular ventas.", severity="error")
            return
        try:
            table = self.query_one(DataTable)
            if table.cursor_row is not None:
                trans_id = int(table.get_cell_at((table.cursor_row, 5)))
                status = table.get_cell_at((table.cursor_row, 4))
                if status == "[ANULADA]":
                    self.app.notify("Esta venta ya fue anulada.", severity="warning")
                    return
                
                def confirm_void(reason):
                    if reason:
                        success, msg = database.void_sale(trans_id, self.app.user_id, reason)
                        if success:
                            self.app.notify(msg, severity="success")
                            self.refresh_history()
                        else:
                            self.app.notify(f"Error: {msg}", severity="error")
                self.app.push_screen(VoidReasonDialog(), confirm_void)
        except Exception:
            self.app.notify("Seleccione una venta para anular.", severity="error")

    def refresh_list(self):
        self.refresh_history()

class EditSaleScreen(Screen):
    """Pantalla para editar una venta ya realizada."""
    def __init__(self, transaction_id):
        super().__init__()
        self.trans_id = transaction_id
        self.cart = {}
        self.total = 0.0
        self.discount_percentage = 0
        self.orig_pay_data = {}

    def compose(self) -> ComposeResult:
        yield Header()
        yield Label(f"EDITAR VENTA ID: {self.trans_id}", id="section_title")
        with Horizontal():
            with Vertical(id="pos_left"):
                yield Label("Añadir/Buscar más productos:")
                yield Input(id="edit_pos_search", placeholder="Código o nombre...")
                yield DataTable(id="edit_pos_table")
            with Vertical(id="pos_right"):
                yield Label("RESUMEN DE EDICIÓN", classes="section_title")
                yield Static("SUBTOTAL: S/ 0.00", id="edit_subtotal_label")
                yield Static("TOTAL: S/ 0.00", id="edit_total_label")
                yield Button("Guardar Cambios (F2)", variant="success", id="btn_save_edit")
                yield Button("Eliminar Item", variant="error", id="btn_del_item_edit")
                yield Button("Cancelar", variant="error", id="btn_cancel_edit")
        yield Footer()

    def on_mount(self):
        table = self.query_one("#edit_pos_table")
        table.add_columns("Código", "Producto", "Cant.", "P.Unit", "Subtotal")
        table.cursor_type = "row"
        
        # Cargar datos actuales de la venta
        sale, items = database.get_sale_full_details(self.trans_id)
        if sale:
            self.orig_pay_data = dict(sale)
            for item in items:
                # Estructura carrito: [nombre, qty, price, stock, marca, unidad]
                self.cart[item['codigo']] = [
                    item['nombre'], item['quantity'], item['unit_price'], 
                    item['stock'], item['fabricante'], item['unidad']
                ]
            self.refresh_edit_table()

    def on_input_submitted(self, event: Input.Submitted):
        if event.input.id == "edit_pos_search":
            term = event.value.strip()
            if term:
                # Reutilizamos lógica de búsqueda del POS (handle_search se puede mover a un mixin o helper si fuera necesario, pero por ahora lo duplicamos simplificado)
                prod = database.get_product(term) or database.get_product_by_name(term)
                if prod:
                    self.app.push_screen(QuantityDialog(prod), lambda qty: self.add_to_edit_cart(prod, qty) if qty else None)
                event.input.value = ""

    def add_to_edit_cart(self, prod, qty):
        code = prod['codigo']
        if code in self.cart: self.cart[code][1] += qty
        else: self.cart[code] = [prod['nombre'], qty, prod['precio_venta'], prod['stock'], prod['fabricante'], prod['unidad']]
        self.refresh_edit_table()

    def refresh_edit_table(self):
        table = self.query_one("#edit_pos_table")
        table.clear()
        subtotal = 0.0
        for code, data in self.cart.items():
            item_sub = data[1] * data[2]
            subtotal += item_sub
            table.add_row(code, data[0], str(data[1]), f"S/ {data[2]:.2f}", f"S/ {item_sub:.2f}")
        
        self.total = subtotal
        self.query_one("#edit_subtotal_label").update(f"SUBTOTAL: S/ {subtotal:.2f}")
        self.query_one("#edit_total_label").update(f"TOTAL: S/ {self.total:.2f}")

    def on_button_pressed(self, event: Button.Pressed):
        if event.button.id == "btn_save_edit":
            self.app.push_screen(PaymentDialog(self.total), self.save_changes)
        elif event.button.id == "btn_del_item_edit":
            try:
                table = self.query_one("#edit_pos_table")
                code = table.get_cell_at((table.cursor_row, 0))
                if code in self.cart:
                    del self.cart[code]
                    self.refresh_edit_table()
            except Exception: pass
        elif event.button.id == "btn_cancel_edit":
            self.dismiss(False)

    def save_changes(self, new_pay_data):
        if new_pay_data:
            items = [(code, data[1], data[2]) for code, data in self.cart.items()]
            try:
                database.update_sale(self.trans_id, self.total, items, new_pay_data)
                self.app.notify("Venta actualizada y stock recalculado.", severity="success")
                self.dismiss(True)
            except Exception as e:
                self.app.notify(f"Error al actualizar: {e}")

class ViewReceiptDialog(Screen):
    """Ventana para visualizar un recibo guardado en el historial."""
    BINDINGS = [Binding("escape", "dismiss", "Cerrar")]

    def __init__(self, transaction_id):
        super().__init__()
        self.trans_id = transaction_id
        self.receipt_content = ""
        self.correlativo = ""

    def compose(self) -> ComposeResult:
        with Vertical(classes="modal-container", id="receipt-view-dialog"):
            yield Label("VISUALIZACIÓN DE COMPROBANTE", id="modal_title")
            yield Static(id="receipt_content", classes="receipt-text-area")
            with Horizontal(classes="form-buttons"):
                yield Button("Ticket Texto", variant="success", id="btn_print_receipt")
                yield Button("Ticket HTML", variant="success", id="btn_print_html_view")
                yield Button("Cerrar (ESC)", variant="error", id="btn_close_view")

    def on_mount(self):
        sale_row, items_row = database.get_sale_full_details(self.trans_id)
        if sale_row:
            self.sale_dict = dict(sale_row)
            self.items_list = [dict(i) for i in items_row]
            self.correlativo = self.sale_dict.get('correlativo', f"ID_{self.trans_id}")
            self.receipt_content = self.formatear_recibo(self.sale_dict, self.items_list)
            self.query_one("#receipt_content").update(self.receipt_content)

    def formatear_recibo(self, sale, items):
        width = 42
        t = f"{'='*width}\n{'FERRETERIA ROLIK':^{width}}\n"
        t += f"{'RUC: 10440809320':^{width}}\n"
        t += f"{'CEL: 988352912 / 932326764':^{width}}\n"
        t += f"{'MZ A LT 26 P.I. MADERA KM 15.5':^{width}}\n"
        t += f"{'='*width}\n"
        t += f"COMPROBANTE: {sale['tipo_comprobante']}\n"
        t += f"NUMERO:      {sale['correlativo']}\n"
        t += f"FECHA:       {sale['date']}\n{'-'*width}\n"
        t += f"{'CANT':<5} {'DESCRIPCION':<24} {'TOTAL':>10}\n{'-'*width}\n"
        for item in items:
            sub = item['quantity'] * item['unit_price']
            desc = f"{item['nombre'][:20]} ({str(item.get('fabricante') or '')[:3]})"
            t += f"{item['quantity']:<5} {desc:<24} {sub:>10.2f}\n"
        t += f"{'-'*width}\n{'TOTAL PAGADO:':<28} S/ {sale['total']:>10.2f}\n"
        t += f"METODO PAGO: {sale['metodo_pago']}\n{'='*width}\n"
        return t

    def on_button_pressed(self, event: Button.Pressed):
        if event.button.id == "btn_print_receipt":
            self.imprimir_recibo_archivo()
        elif event.button.id == "btn_print_html_view":
            self.imprimir_ticket_html_reprint()
        else:
            self.dismiss()

    def imprimir_ticket_html_reprint(self):
        """Genera y abre el HTML para un ticket antiguo."""
        try:
            os.makedirs("ventas", exist_ok=True)
            filename = os.path.join("ventas", f"Ticket_{self.correlativo}.html")
            items_html = ""
            for item in self.items_list:
                items_html += f"<tr><td>{item['quantity']}</td><td>{item['nombre'][:25]}</td><td style='text-align: right;'>{(item['quantity']*item['unit_price']):.2f}</td></tr>"

            subtotal = self.sale_dict['total'] / 1.18
            igv = self.sale_dict['total'] - subtotal
            
            html = f"""
            <!DOCTYPE html><html><head><meta charset="UTF-8"><style>@page {{ margin: 0; }} body {{ width: 72mm; font-family: 'Arial Narrow', Arial, sans-serif; font-size: 12px; margin: 0; padding: 4mm; color: black; }} .center {{ text-align: center; }} .bold {{ font-weight: bold; }} .line {{ border-top: 1px dashed black; margin: 5px 0; }} table {{ width: 100%; border-collapse: collapse; margin: 10px 0; }} th {{ border-bottom: 1px solid black; text-align: left; }} .total-table {{ margin-top: 10px; float: right; width: 60%; }} .footer {{ margin-top: 20px; font-size: 10px; }}</style></head><body onload="window.print();">
            <div class="center"><div class="bold" style="font-size: 16px;">FERRETERÍA ROLIK</div><div>RUC: 10440809320</div><div>Mz A lt 26 P.I. Madera KM 15.5</div><div>Cel: 988352912 / 932326764</div><div class="line"></div><div class="bold">{self.sale_dict['tipo_comprobante']}</div><div>{self.correlativo}</div><div>Fecha: {self.sale_dict['date']}</div></div>
            <div style="margin-top: 10px;"><div><b>Cliente:</b> {self.sale_dict.get('cliente_nombre', 'PÚBLICO GENERAL')}</div><div><b>Doc:</b> {self.sale_dict.get('cliente_documento', '00000000')}</div></div>
            <table><thead><tr><th style="width: 15%;">Cant</th><th style="width: 60%;">Descrip</th><th style="width: 25%; text-align: right;">Total</th></tr></thead><tbody>{items_html}</tbody></table>
            <div class="line"></div><table class="total-table"><tr><td>SUBTOTAL:</td><td style="text-align: right;">S/ {subtotal:.2f}</td></tr><tr><td>IGV (18%):</td><td style="text-align: right;">S/ {igv:.2f}</td></tr><tr class="bold" style="font-size: 14px;"><td>TOTAL:</td><td style="text-align: right;">S/ {self.sale_dict['total']:.2f}</td></tr></table>
            <div style="clear: both; margin-top: 10px;"><div><b>Pago:</b> {self.sale_dict['metodo_pago']}</div><div><b>Recibido:</b> S/ {self.sale_dict.get('monto_pagado', 0):.2f}</div><div><b>Vuelto:</b> S/ {self.sale_dict.get('vuelto', 0):.2f}</div></div>
            <div class="center footer"><div class="line"></div></div></body></html>
            """
            with open(filename, "w", encoding="utf-8") as f: f.write(html)
            if os.name == 'nt': os.startfile(filename)
        except Exception as e: self.app.notify(f"Error HTML: {e}", severity="error")

    def imprimir_recibo_archivo(self):
        """Guarda el recibo actual en un archivo .txt."""
        if not self.receipt_content:
            self.app.notify("No hay contenido de recibo para imprimir.", severity="error")
            return
        try:
            os.makedirs("ventas", exist_ok=True)
            filename = os.path.join("ventas", f"Ticket_{self.correlativo}.txt")
            with open(filename, "w", encoding="utf-8") as f:
                f.write(self.receipt_content)
            
            self.app.notify(f"Recibo guardado: {filename}", severity="information")
            if os.name == 'nt':
                os.startfile(filename)
        except Exception as e:
            self.app.notify(f"Error al imprimir: {e}", severity="error")

class CommissionReportScreen(Screen):
    BINDINGS = [Binding("escape", "app.pop_screen", "Volver"), ("r", "refresh_report", "Refrescar")]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.selected_user_id = None
        self.start_date = ""
        self.end_date = ""

    def compose(self) -> ComposeResult:
        yield Header()
        yield Label("REPORTE DE COMISIONES", id="section_title")
        with Horizontal(id="report_filters"):
            yield Label("Vendedor:")
            yield Input(placeholder="Todos", id="filter_seller_name", classes="filter-input")
            yield Button("Seleccionar Vendedor", id="btn_select_seller")
            yield Label("Desde (YYYY-MM-DD):")
            yield Input(placeholder="YYYY-MM-DD", id="filter_start_date", classes="filter-input")
            yield Label("Hasta (YYYY-MM-DD):")
            yield Input(placeholder="YYYY-MM-DD", id="filter_end_date", classes="filter-input")
            yield Button("Refrescar (R)", id="btn_refresh_report")
        yield DataTable(id="commissions_report_table")
        yield Footer()

    def on_mount(self) -> None:
        table = self.query_one(DataTable)
        table.cursor_type = "row"
        table.add_columns("ID Comisión", "Fecha", "Monto Comisión", "Vendedor", "ID Venta", "Total Venta")
        self.refresh_report()

    def refresh_report(self):
        try:
            start_date = self.query_one("#filter_start_date", Input).value
            end_date = self.query_one("#filter_end_date", Input).value
            
            table = self.query_one(DataTable)
            table.clear()

            commissions = database.get_commissions_history(
                user_id=self.selected_user_id,
                start_date=start_date if start_date else None,
                end_date=end_date if end_date else None
            )

            total_commissions_amount = 0.0
            for comm in commissions:
                table.add_row(
                    str(comm['commission_id']),
                    str(comm['commission_date']),
                    f"${comm['commission_amount']:.2f}",
                    str(comm['seller_name']),
                    str(comm['transaction_id']),
                    f"${comm['transaction_total']:.2f}"
                )
                total_commissions_amount += comm['commission_amount']
            
            table.add_row("", "", "", "", "", "", key="total-row", classes="footer")
            table.add_row("TOTAL COMISIONES:", "", "", "", "", f"${total_commissions_amount:.2f}", key="total-row-val", classes="footer")

        except Exception as e:
            self.app.notify(f"Error al cargar el reporte de comisiones: {e}", severity="error")

    def action_refresh_report(self):
        self.refresh_report()

    def on_button_pressed(self, event: Button.Pressed) -> None:
        if event.button.id == "btn_refresh_report":
            self.refresh_report()
        elif event.button.id == "btn_select_seller":
            self.app.push_screen(UserSelectionDialog(), callback=self.on_seller_selected)
        elif event.button.id == "btn_back":
            self.app.pop_screen()

    def on_seller_selected(self, user_data: dict | None):
        if user_data:
            self.selected_user_id = user_data['id']
            self.query_one("#filter_seller_name", Input).value = user_data['username']
        else:
            self.selected_user_id = None
            self.query_one("#filter_seller_name", Input).value = "Todos"
        self.refresh_report()

class PurchaseOrderListScreen(Screen):
    BINDINGS = [
        Binding("escape", "app.pop_screen", "Volver"),
        Binding("n", "create_po", "Nuevo Pedido"),
        Binding("r", "receive_po", "Recibir Pedido"),
        Binding("v", "view_po", "Ver Detalle"),
        Binding("s", "change_status", "Cambiar Estado")
    ]
    
    def compose(self) -> ComposeResult:
        yield Header()
        yield Vertical(
            Label("ROLIK - ÓRDENES DE COMPRA", id="section_title"),
            DataTable(id="po_table"),
            Horizontal(
                Button("Nuevo (N)", variant="success", id="btn_new"), 
                Button("Ver Detalle (V)", variant="primary", id="btn_view"),
                Button("Cambiar Estado (S)", variant="warning", id="btn_status"),
                Button("Recibir (R)", variant="warning", id="btn_receive"),
                Button("Volver (ESC)", variant="error", id="btn_back"), 
                id="po_actions"))
        yield Footer()
    
    def on_mount(self):
        table = self.query_one(DataTable)
        table.cursor_type = "row"
        table.add_columns("ID", "NÚMERO OC", "PROVEEDOR", "FECHA PEDIDO", "FECHA EST.", "ESTADO", "TOTAL")
        self.refresh_pos()
    
    def refresh_pos(self):
        table = self.query_one(DataTable)
        table.clear()
        for order in database.get_all_purchase_orders():
            # numero_oc es el nuevo campo robusto
            table.add_row(
                str(order['id']), 
                str(order['numero_oc'] or "N/A"),
                order['proveedor'], 
                str(order['fecha_pedido'])[:10],
                str(order['fecha_estimada'] or "N/A"),
                order['estado'],
                f"${order['total']:.2f}"
            )
    
    def action_view_po(self):
        try:
            table = self.query_one(DataTable)
            order_id = int(table.get_cell_at((table.cursor_row, 0)))
            self.app.push_screen(ViewPODetailsScreen(order_id))
        except Exception:
            self.app.notify("Selecciona una orden para ver detalles.", severity="error")

    def action_change_status(self):
        try:
            table = self.query_one(DataTable)
            order_id = int(table.get_cell_at((table.cursor_row, 0)))
            self.app.push_screen(ChangeStatusDialog(order_id), lambda success: self.refresh_pos() if success else None)
        except Exception:
            self.app.notify("Selecciona una orden para cambiar su estado.", severity="error")
    
    def on_po_created(self, success: bool):
        if success: self.refresh_pos(); self.app.notify("Orden de compra creada.", severity="success")
    
    def action_create_po(self): self.app.push_screen(CreatePurchaseOrderScreen(), self.on_po_created)
    
    def action_receive_po(self):
        try:
            table = self.query_one(DataTable); order_id = table.get_cell_at((table.cursor_row, 0))
            estado = table.get_cell_at((table.cursor_row, 3))
            if estado == "RECIBIDO": self.app.notify("Esta orden ya ha sido recibida.", severity="warning"); return
            database.receive_purchase_order(order_id)
            self.app.notify(f"Orden #{order_id} recibida. Stock actualizado.", severity="success"); self.refresh_pos()
            self.app.query_one(InventoryScreen).refresh_inventory() # Actualizar inventario si está montado
        except Exception: self.app.notify("Selecciona una orden de la tabla para recibir.", severity="error")
    
    def on_button_pressed(self, event: Button.Pressed):
        if event.button.id == "btn_new": self.action_create_po()
        elif event.button.id == "btn_receive": self.action_receive_po()
        elif event.button.id == "btn_back": self.app.pop_screen()

class ChangeStatusDialog(Screen):
    def __init__(self, order_id):
        super().__init__(); self.order_id = order_id
    
    def compose(self) -> ComposeResult:
        order = database.get_purchase_order_by_id(self.order_id)
        with Vertical(classes="modal-container"):
            yield Label(f"CAMBIAR ESTADO - ORDEN {order['numero_oc']}")
            if order['estado'] in ["RECIBIDA", "RECIBIDO"]:
                yield Label("ESTADO BLOQUEADO: La orden ya fue RECIBIDA.", classes="error")
                yield Button("Cerrar", id="btn_cancel")
            else:
                yield Button("PENDIENTE", id="PENDIENTE")
                yield Button("APROBADA", id="APROBADA")
                yield Button("ENVIADA", id="ENVIADA")
                yield Button("RECIBIDA (ACTUALIZA STOCK)", id="RECIBIDA", variant="warning")
                yield Button("CANCELADA", id="CANCELADA", variant="error")
                yield Button("Cancelar", id="btn_cancel")

    def on_button_pressed(self, event: Button.Pressed):
        if event.button.id == "btn_cancel":
            self.dismiss(False)
        else:
            database.update_purchase_order_status(self.order_id, event.button.id)
            self.app.notify(f"Estado actualizado a {event.button.id}")
            self.dismiss(True)

class ViewPODetailsScreen(Screen):
    def __init__(self, order_id):
        super().__init__(); self.order_id = order_id
        self.order_data = None
    
    def compose(self) -> ComposeResult:
        yield Header()
        yield Label(f"DETALLE DE ORDEN DE COMPRA", id="section_title")
        yield Static(id="po_header_info", classes="po-details-text")
        yield DataTable(id="po_details_table")
        yield Static(id="po_footer_info", classes="po-details-text")
        with Horizontal(classes="form-buttons"):
            yield Button("Exportar TXT", variant="primary", id="btn_export_txt")
            yield Button("Exportar PDF", variant="success", id="btn_export_pdf")
            yield Button("Cerrar (ESC)", variant="error", id="btn_close")
        yield Footer()

    def on_mount(self):
        order = database.get_purchase_order_by_id(self.order_id)
        self.order_data = order
        details = database.get_purchase_order_details(self.order_id)
        
        self.query_one("#section_title").update(f"DETALLE DE ORDEN: {order['numero_oc']}")
        
        header = f"""
NÚMERO OC: {order['numero_oc']} | FECHA: {order['fecha_pedido']}
PROVEEDOR: {order['proveedor_nombre']} | RUC: {order['ruc_dni'] or 'N/A'}
DIRECCIÓN: {order['direccion'] or 'N/A'} | TEL: {order['telefono'] or 'N/A'}
ESTADO: {order['estado']} | PAGO: {order['condicion_pago'] or 'N/A'}
FECHA ESTIMADA: {order['fecha_estimada'] or 'N/A'} | LUGAR: {order['lugar_entrega'] or 'N/A'}
"""
        self.query_one("#po_header_info").update(header)
        
        table = self.query_one("#po_details_table")
        table.add_columns("Producto", "Cantidad", "P. Unit", "Subtotal")
        for item in details:
            sub = item['cantidad'] * item['precio_compra_unitario']
            table.add_row(item['nombre'], str(item['cantidad']), f"${item['precio_compra_unitario']:.2f}", f"${sub:.2f}")
        
        footer = f"""
SUBTOTAL: ${order['subtotal']:.2f} | IGV (18%): ${order['igv']:.2f} | TOTAL: ${order['total']:.2f}
"""
        self.query_one("#po_footer_info").update(footer)

    def on_button_pressed(self, event: Button.Pressed):
        if event.button.id == "btn_export_txt":
            self.export_to_txt()
        elif event.button.id == "btn_export_pdf":
            self.export_to_pdf()
        elif event.button.id == "btn_close":
            self.app.pop_screen()

    def export_to_txt(self):
        try:
            o = self.order_data
            details = database.get_purchase_order_details(self.order_id)
            filename = f"Orden_{o['numero_oc']}.txt"
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(f"FERRETERIA ROLIK\n")
                f.write(f"RUC: 10440809320 | MZ A LT 26 P.I. MADERA\n")
                f.write("-" * 50 + "\n")
                f.write(f"ORDEN DE COMPRA: {o['numero_oc']}\n")
                f.write(f"Fecha: {o['fecha_pedido']}\n")
                f.write(f"Proveedor: {o['proveedor_nombre']} | RUC: {o['ruc_dni']}\n")
                f.write("-" * 50 + "\n")
                for d in details:
                    sub = d['cantidad'] * d['precio_compra_unitario']
                    f.write(f"{d['nombre'][:30]:<30} | {d['cantidad']:<5} | ${d['precio_compra_unitario']:>8.2f} | ${sub:>8.2f}\n")
                f.write("-" * 50 + "\n")
                f.write(f"SUBTOTAL: ${o['subtotal']:>38.2f}\n")
                f.write(f"IGV (18%): ${o['igv']:>38.2f}\n")
                f.write(f"TOTAL: ${o['total']:>41.2f}\n")
                f.write(f"Estado: {o['estado']}\n")
            self.app.notify(f"Archivo TXT generado: {filename}", severity="information")
        except Exception as e:
            self.app.notify(f"Error al exportar TXT: {e}", severity="error")

    def export_to_pdf(self):
        try:
            order = self.order_data
            details = database.get_purchase_order_details(self.order_id)
            
            pdf = FPDF()
            pdf.add_page()
            
            # Comprador
            pdf.set_font("Arial", "B", 12)
            pdf.cell(0, 10, "FERRETERIA ROLIK", ln=True)
            pdf.set_font("Arial", "", 10)
            pdf.cell(0, 5, "RUC: 10440809320 | Cel: 988352912/932326764", ln=True)
            pdf.cell(0, 5, "Mz A lt 26 Parque industrial de la madera km 15.5", ln=True)
            pdf.ln(5)
            
            # Encabezado
            pdf.set_font("Arial", "B", 16)
            pdf.cell(0, 10, f"ORDEN DE COMPRA {order['numero_oc']}", ln=True, align="C")
            pdf.ln(5)
            
            # Datos del Proveedor y Orden
            pdf.set_font("Arial", "", 10)
            pdf.cell(0, 7, f"Proveedor: {order['proveedor_nombre']}", ln=True)
            pdf.cell(0, 7, f"RUC/DNI: {order['ruc_dni'] or 'N/A'}", ln=True)
            pdf.cell(0, 7, f"Direccion: {order['direccion'] or 'N/A'}", ln=True)
            pdf.cell(0, 7, f"Telefono: {order['telefono'] or 'N/A'}", ln=True)
            pdf.ln(3)
            pdf.cell(0, 7, f"Fecha Pedido: {order['fecha_pedido']}", ln=True)
            pdf.cell(0, 7, f"Condicion Pago: {order['condicion_pago'] or 'N/A'}", ln=True)
            pdf.cell(0, 7, f"Lugar Entrega: {order['lugar_entrega'] or 'N/A'}", ln=True)
            pdf.ln(10)
            
            # Tabla de Productos
            pdf.set_font("Arial", "B", 10)
            pdf.set_fill_color(200, 220, 255)
            pdf.cell(80, 10, "Producto", border=1, fill=True)
            pdf.cell(30, 10, "Cant.", border=1, fill=True, align="C")
            pdf.cell(40, 10, "Precio Unit.", border=1, fill=True, align="C")
            pdf.cell(40, 10, "Subtotal", border=1, fill=True, align="C")
            pdf.ln()
            
            pdf.set_font("Arial", "", 10)
            for item in details:
                sub = item['cantidad'] * item['precio_compra_unitario']
                pdf.cell(80, 8, item['nombre'][:40], border=1)
                pdf.cell(30, 8, str(item['cantidad']), border=1, align="C")
                pdf.cell(40, 8, f"${item['precio_compra_unitario']:.2f}", border=1, align="C")
                pdf.cell(40, 8, f"${sub:.2f}", border=1, align="C")
                pdf.ln()
            
            # Totales
            pdf.ln(5)
            pdf.set_font("Arial", "B", 11)
            pdf.cell(150, 8, "SUBTOTAL:", align="R"); pdf.cell(40, 8, f"${order['subtotal']:.2f}", border=1, align="C"); pdf.ln()
            pdf.cell(150, 8, "IGV (18%):", align="R"); pdf.cell(40, 8, f"${order['igv']:.2f}", border=1, align="C"); pdf.ln()
            pdf.set_fill_color(255, 230, 150)
            pdf.cell(150, 8, "TOTAL FINAL:", align="R"); pdf.cell(40, 8, f"${order['total']:.2f}", border=1, fill=True, align="C"); pdf.ln()
            
            filename = f"Orden_Compra_{order['numero_oc']}.pdf"
            pdf.output(filename)
            self.app.notify(f"PDF generado: {filename}", severity="information")
            if os.name == 'nt': os.startfile(filename)
        except Exception as e:
            self.app.notify(f"Error al exportar PDF: {e}", severity="error")

class CreatePurchaseOrderScreen(Screen):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs); self.po_items = {}
    
    def compose(self) -> ComposeResult:
        yield Header()
        with VerticalScroll(id="po-form-scroll"):
            yield Label("NUEVA ORDEN DE COMPRA", id="modal_title")
            
            with Horizontal(classes="po-row"):
                with Vertical(classes="po-col"):
                    yield Label("Proveedor:"); yield Input(id="in_proveedor_po")
                    yield Label("RUC Proveedor (11 dígitos):"); yield Input(id="in_ruc_po", restrict=r"[0-9]*")
                    yield Label("Fecha Est. Entrega:"); yield Input(placeholder="YYYY-MM-DD", id="in_po_delivery_date")
                with Vertical(classes="po-col"):
                    yield Label("Condición Pago:"); yield Input(placeholder="Contado/Crédito", id="in_po_payment")
                    yield Label("Lugar Entrega:"); yield Input(id="in_po_place")
            
            yield Label("AÑADIR PRODUCTOS", classes="sub-title")
            with Horizontal(id="po-product-adder"):
                yield Input(placeholder="Código o Nombre del Producto", id="in_po_search", classes="input-search")
                yield Input(placeholder="Cant.", id="in_po_qty", restrict=r"[0-9]*", classes="input-small")
                yield Input(placeholder="Costo U.", id="in_po_cost", restrict=r"[0-9.]*", classes="input-small")
                yield Button("Añadir", variant="primary", id="btn_add_item")
            
            yield DataTable(id="po_items_table")
            
            with Horizontal(classes="form-buttons-po"):
                yield Button("CREAR ORDEN", variant="success", id="btn_create_po")
                yield Button("CANCELAR", variant="error", id="btn_cancel_po")
        yield Footer()
    
    def on_mount(self): 
        self.query_one("#po_items_table").add_columns("Código", "Cant.", "Costo", "Producto")
    
    def on_button_pressed(self, event: Button.Pressed):
        if event.button.id == "btn_add_item":
            search_term = self.query_one("#in_po_search", Input).value.strip()
            qty_str = self.query_one("#in_po_qty", Input).value.strip()
            cost_str = self.query_one("#in_po_cost", Input).value.strip()
            
            if not (search_term and qty_str and cost_str):
                self.app.notify("Búsqueda, Cantidad y Costo son requeridos.", severity="error")
                return
            
            try:
                qty = int(qty_str)
                cost = float(cost_str)
                if qty <= 0 or cost <= 0:
                    self.app.notify("Cantidad y costo deben ser mayores a cero.", severity="error")
                    return
            except ValueError:
                self.app.notify("Cantidad (entero) y Costo (decimal) deben ser números válidos.", severity="error")
                return
            
            # 1. Intentar buscar por código
            product = database.get_product(search_term)
            
            # 2. Si no se encuentra por código, intentar por nombre
            if not product:
                product = database.get_product_by_name(search_term)
            
            if not product:
                self.app.notify(f"Producto '{search_term}' no encontrado (por código ni nombre).", severity="error")
                return
            
            code = product['codigo']
            # Añadir al diccionario local y refrescar tabla
            self.po_items[code] = [product['nombre'], qty, cost]
            self.refresh_po_items_table()
            
            # Limpiar campos de producto para el siguiente
            self.query_one("#in_po_search", Input).value = ""
            self.query_one("#in_po_qty", Input).value = ""
            self.query_one("#in_po_cost", Input).value = ""
            self.query_one("#in_po_search", Input).focus()
            
        elif event.button.id == "btn_create_po":
            proveedor = self.query_one("#in_proveedor_po").value
            ruc = self.query_one("#in_ruc_po").value
            if not proveedor or not ruc or not self.po_items:
                self.app.notify("Proveedor, RUC y al menos un producto son requeridos.", severity="error")
                return
            
            if len(ruc) != 11:
                self.app.notify("El RUC debe tener 11 dígitos.", severity="error")
                return
            
            po_data = {
                'fecha_estimada': self.query_one("#in_po_delivery_date").value,
                'condicion_pago': self.query_one("#in_po_payment").value,
                'lugar_entrega': self.query_one("#in_po_place").value,
            }
            
            items_for_db = [(code, data[1], data[2]) for code, data in self.po_items.items()]
            try:
                database.create_purchase_order(proveedor, ruc, items_for_db, po_data)
                self.dismiss(True)
            except Exception as e:
                self.app.notify(f"Error: {e}", severity="error")
        elif event.button.id == "btn_cancel_po": self.dismiss(False)
    
    def refresh_po_items_table(self):
        table = self.query_one("#po_items_table"); table.clear()
        for code, (nombre, qty, cost) in self.po_items.items(): table.add_row(code, str(qty), f"${cost:.2f}", nombre)

class ERPApp(App):
    TITLE = "Ferretería ROLIK - Sistema de Gestión"
    CSS = """
    Screen { background: #1a1b26; color: #c0caf5; }
    #menu_container { align: center middle; height: 100%; }
    #buttons_grid {
        layout: grid;
        grid-size: 3;
        grid-gutter: 1;
        padding: 1;
        align: center middle;
        width: 100%;
        height: auto;
    }
    #buttons_grid Button {
        width: 100%;
        height: 3;
        text-style: bold;
        margin: 0;
    }
    #title { text-align: center; width: 100%; padding: 1; margin-bottom: 1; background: #7aa2f7; color: #1a1b26; text-style: bold; }
    #section_title { text-align: center; width: 100%; padding: 1; background: #bb9af7; color: #1a1b26; text-style: bold; }
    Button { width: 35; height: 3; text-style: bold; }
    #inventory_actions Button, #po_actions Button, #user_actions Button, #customer_actions Button {
        width: auto;
        min-width: 16;
        height: 3;
        margin: 0 1;
    }
    #inventory_actions, #po_actions, #user_actions, #report_actions { height: auto; padding: 1; align: center middle; }
    #search_sort_bar {
        align: center middle;
        padding: 0 1;
        height: 5;
    }
    #search_box {
        width: 40%;
    }
    #search_sort_bar Button {
        width: auto;
        min-width: 15;
    }
    DataTable { height: 70%; margin: 1; border: double #7aa2f7; }
    .low-stock { background: #b48608 40%; }
    .no-stock { background: #c10000 50%; text-style: bold; }
    
    #form-body {
        padding: 0 2;
    }
    #modal_title { 
        text-align: center; 
        width: 100%; 
        padding-top: 1;
        text-style: bold; 
    }
    .form-buttons { 
        align: center middle; 
        height: 5;
        dock: bottom;
    }
    Input { margin-bottom: 1; } 
    Label { margin-top: 1; }
    #pos_left { width: 65%; padding: 0 1; } #pos_right { width: 35%; padding: 0 1; border-left: double #7aa2f7; }
    #pos_total_label { border: wide #f7768e; text-align: center; text-style: bold; padding: 1; margin: 2 0; }
    #ticket_text { width: auto; } .modal-container { align: center middle; padding: 2; background: #24283b; border: heavy #7aa2f7; width: auto; height: auto; }
    #po_form_left { width: 30%; } #po_form_right { width: 70%; }
    #user_actions { height: auto; padding: 1; align: center middle; }
    
    #login-container {
        width: 50;
        height: auto;
        border: heavy #7aa2f7;
    }
    .po-details-text {
        padding: 1;
        margin: 1;
        background: #24283b;
        border: solid #7aa2f7;
        color: #7aa2f7;
        text-style: bold;
    }
    #po_header_form {
        height: auto;
        margin-bottom: 1;
    }
    #po_header_form Vertical {
        padding: 0 1;
    }
    .po-row { height: auto; margin-bottom: 1; }
    .po-col { width: 50%; padding: 0 1; }
    .sub-title { text-align: center; margin: 1 0; text-style: bold; color: #7aa2f7; border-top: solid #7aa2f7; padding-top: 1; }
    #po-product-adder { height: 3; margin-bottom: 1; align: center middle; }
    #po-product-adder .input-search { width: 45%; margin: 0 1; }
    #po-product-adder .input-small { width: 15%; margin: 0 1; }
    #po-product-adder Button { width: 15; margin: 0 1; }
    #po_items_table { height: 12; border: double #bb9af7; margin: 1 0; }
    .form-buttons-po { align: center middle; height: 3; margin-top: 1; }
    .form-buttons-po Button { width: 20; margin: 0 2; }
    #report_menu_container {
        padding: 1 2;
        align: center middle;
    }
    .report-row {
        height: auto;
        margin-bottom: 1;
        align: center middle;
    }
    .report-row Button {
        width: 35;
        margin: 0 2;
    }
    .report-summary-box {
        background: #24283b;
        color: #7aa2f7;
        border: double #7aa2f7;
        padding: 2;
        margin: 2;
        width: 60;
        height: auto;
        text-style: bold;
    }
    #filter_bar {
        height: 5;
        align: center middle;
        padding: 1;
    }
    #filter_bar Input {
        width: 25;
        margin: 0 1;
    }
    #receipt-view-dialog {
        width: 60;
        height: 55;
        border: thick #7aa2f7;
        background: #1a1b26;
    }
    .receipt-text-area {
        background: #24283b;
        color: #c0caf5;
        padding: 1 2;
        border: solid #3b4261;
        margin: 1;
        height: 30;
        text-style: bold;
    }
    .help-text {
        text-align: center;
        color: #bb9af7;
        text-style: italic;
        margin-bottom: 1;
    }
    #client-search-row {
        height: 3;
        margin-bottom: 1;
    }
    #client-search-row Input {
        width: 70%;
    }
    #client-search-row Button {
        width: 25%;
        margin-left: 1;
    }
    #customer_search_bar {
        height: 5;
        align: center middle;
        padding: 1;
    }
    #customer_search_bar Input {
        width: 60%;
        margin-right: 2;
    }
    #customer_search_bar Button {
        width: 20;
    }
    #customer_actions {
        height: auto;
        padding: 1;
        align: center middle;
    }
    #customer_actions Button {
        width: 25;
        margin: 0 1;
    }
    #cust-form-container {
        width: 60;
        height: 40;
        border: thick #7aa2f7;
        background: #1a1b26;
        padding: 1 2;
    }
    #search-dialog {
        width: 80;
        height: 30;
        border: thick #bb9af7;
        background: #1a1b26;
    }
    #pay-dialog-container {
        width: 70;
        height: 55;
        border: thick #7aa2f7;
        background: #1a1b26;
        padding: 1 2;
    }
    .options-row {
        margin: 1 0;
        height: 3;
        align: center middle;
    }
    .options-row Button {
        width: 12;
        margin: 0 1;
        height: 3;
    }
    #search_results_table {
        height: 20;
        border: solid #7aa2f7;
    }
    .opt-btn { width: 12; margin: 0 1; background: #3b4261; }
    .opt-btn:focus { background: #7aa2f7; color: #1a1b26; }
    #lbl_change { margin-top: 1; padding: 1; border: solid #7aa2f7; text-align: center; color: #7aa2f7; text-style: bold; }
    #pos_in_search { width: 100%; margin-bottom: 1; }
    #in_paid_amount { margin-top: 1; }
    #pay-dialog-full {
        width: 110;
        height: 60;
        border: thick #7aa2f7;
        background: #1a1b26;
        padding: 1;
    }
    #pay_top_grid {
        layout: grid;
        grid-size: 2;
        height: 18;
        grid-gutter: 2;
        margin-bottom: 1;
    }
    .pay_box {
        border: solid #3b4261;
        padding: 1;
        background: #24283b;
    }
    #pay_bottom_container {
        border: solid #7aa2f7;
        background: #1a1b26;
        padding: 1;
        height: auto;
    }
    #pay_confirm_table {
        height: 14;
        margin: 1 0;
        border: solid #3b4261;
    }
    #pay_extra_search_row {
        height: 3;
        margin: 1 0;
    }
    #in_pay_extra_search {
        width: 100%;
        border: double #bb9af7;
    }
    #pay_edit_controls {
        height: 3;
        align: center middle;
        margin-bottom: 1;
    }
    #pay_edit_controls Button {
        width: 20;
        margin: 0 1;
    }
    #pay_confirm_subtotal {
        text-align: right;
        text-style: bold;
        color: #f7768e;
        background: #24283b;
        padding: 1;
        border-top: double #7aa2f7;
    }
    #receipt-buttons-grid {
        layout: grid;
        grid-size: 2;
        grid-gutter: 1;
        margin: 1 0;
        height: auto;
        width: 100%;
    }
    #receipt-buttons-grid Button {
        width: 100%;
        height: 3;
    }
    #btn_done {
        width: 100%;
        margin-top: 1;
    }
    #final-receipt-container {
        width: 65;
        height: auto;
        padding: 1 2;
    }
    """
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.user_id = None
        self.user_role = None
        self.user_name = None
        self.is_discount_enabled = False
        self.max_discount_percentage = 0
        self.commission_rate = 0.0
        self.permissions = set()

    def on_mount(self) -> None: 
        database.init_db()
        self.show_login()

    def show_login(self):
        self.push_screen(LoginScreen(), callback=self.on_login_finished)

    def has_permission(self, permission_name: str) -> bool:
        """Comprueba si el usuario logueado tiene un permiso específico."""
        return permission_name in self.permissions

    def on_login_finished(self, user_data: dict | None):
        if user_data:
            self.user_id = user_data['id']
            self.user_role = user_data['role']
            self.user_name = user_data['username']
            self.is_discount_enabled = user_data.get('is_discount_enabled', 0)
            self.max_discount_percentage = user_data.get('max_discount_percentage', 0)
            self.commission_rate = user_data.get('commission_rate', 0.0)
            self.permissions = user_data.get('permissions', set()) # Cargar permisos
            self.notify(f"Bienvenido {self.user_name} ({self.user_role})", severity="success")
            self.push_screen(MainMenu())
        else:
            self.exit()

if __name__ == "__main__":
    app = ERPApp()
    app.run()
